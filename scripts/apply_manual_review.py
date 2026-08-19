#!/usr/bin/env python3
"""Apply human-reviewed identity decisions back into enriched.json.

Fills the gap batch_enrich.py leaves open: for records batch_enrich.py
flagged as needing review (ambiguous / not_found / possible_move /
needs_review_contaminated / needs_review_field_mismatch / api_error), a
human can open the merged Excel, look at OpenAlex directly (or a school
homepage, or Semantic Scholar — anywhere, as long as a human is the one
judging identity, not this script), and write their conclusion into the
"人工核实结果" column merge_to_excel.py adds for exactly this purpose. This
script reads that column back, re-fetches the confirmed person's full
profile the same way batch_enrich.py does for an automatic match, and
writes an updated enriched.json — so a manually-confirmed record ends up
with the exact same shape (works, topics, affiliations, ...) as an
automatically-resolved one, and can flow through confidence.py /
journal_ranking.py / merge_to_excel.py identically afterward.

Accepted values for the "人工核实结果" column, one per row (case-insensitive,
leading/trailing whitespace ignored):
  <empty>              not reviewed yet — record is left untouched
  A1234567890           an OpenAlex author ID — "this specific person is who
                        it actually is"; re-fetched fresh regardless of what
                        batch_enrich.py originally found
  ok                    "the single candidate batch_enrich.py already listed
                        under OA_候选人(需人工确认) is correct" — only valid
                        when the record has exactly one candidate listed
  skip                  "confirmed: not findable on OpenAlex, stop asking" —
                        marks the record so it won't be flagged again, but
                        does NOT fabricate any OpenAlex data for it
  anything else          treated as a formatting mistake, left untouched,
                        and listed in the "无法识别" summary at the end —
                        never silently guessed at

Usage:
  python apply_manual_review.py 结果_OA补全_已核实.xlsx enriched.json \
      --sheet 全部教授_论文与聚焦度 --mailto you@x.com \
      --window-a-since 2020 --out enriched_reviewed.json
"""
import argparse, json, re, sys
import openpyxl

import fetch_openalex as oa

OPENALEX_ID_RE = re.compile(r"^A\d{5,}$", re.I)


def _norm(s):
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


def read_review_column(xlsx_path, sheet, name_col, inst_col):
    """Returns {(normalized_name, normalized_institution): raw_review_text}
    for every row where the "人工核实结果" column is non-empty. Matching
    enriched.json records by (name, institution) rather than row position,
    since a user may have sorted/filtered the sheet after downloading it —
    row position is not a safe join key, the original roster columns are.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet] if sheet in wb.sheetnames else wb.active
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        review_col = header.index("人工核实结果") + 1
    except ValueError:
        sys.exit("错误：这份表里没有找到「人工核实结果」这一列——请用 merge_to_excel.py "
                 "重新生成过的 Excel（这一列是自动加的），不要用手动整理过的旧表格。")

    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, inst = row[name_col - 1], row[inst_col - 1]
        review = row[review_col - 1]
        if not name or not review or not str(review).strip():
            continue
        out[(_norm(name), _norm(inst))] = str(review).strip()
    return out


def build_add(aid, mailto, window_a_since, match_method):
    """Same field construction batch_enrich.py uses for an automatic 'ok'
    match — kept in one place so a manually-confirmed record and an
    automatically-resolved one are indistinguishable in shape afterward."""
    prof = oa.profile(aid, mailto)
    works = oa.works(aid, mailto, window_a_since)
    recent = [w for w in works if (w.get("year") or 0) >= (window_a_since or 0)]
    return {
        "openalex_id": aid,
        "match_method": match_method,
        "orcid": prof.get("orcid"),
        "academic_start_year": prof.get("earliest_affiliation_year"),
        "last_known_institutions": prof.get("last_known_institutions"),
        "effective_institutions": prof.get("effective_institutions"),
        "affiliation_institutions": prof.get("affiliation_institutions"),
        "topics": prof.get("topics"),
        "counts_by_year": prof.get("counts_by_year"),
        "cited_by_count": prof.get("cited_by_count"),
        "total_works_count": prof.get("works_count"),
        "recent_works_count": len(recent),
        "recent_works": [{"title": w["title"], "year": w["year"], "venue": w["venue"],
                          "published": w["is_published"], "doi": w["doi"],
                          "url": w.get("url"), "institutions": w.get("institutions")}
                         for w in recent[:12]],
        "works": works,
        "enrich_status": "ok",
        # 人工核实过的记录，把原来那些"为什么需要复核"的痕迹清掉——人已经确认过了，
        # 不该在最终结果里继续显示"疑似身份污染"这类现在已经不适用的旧标记。
        "candidates": [], "verify_url": None,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", help="merge_to_excel.py 生成、你已经填过「人工核实结果」列的那份Excel")
    ap.add_argument("enriched", help="对应的原始 enriched.json（或 ranked.json）")
    ap.add_argument("--sheet", required=True)
    ap.add_argument("--name-col", type=int, default=4, help="姓名列，默认第4列，和roster.xlsx保持一致")
    ap.add_argument("--inst-col", type=int, default=3, help="机构列，默认第3列")
    ap.add_argument("--mailto", required=True)
    ap.add_argument("--api-key", help="OpenAlex API key (required since Feb 2026 — "
                                      "free at https://openalex.org/settings/api)")
    ap.add_argument("--window-a-since", type=int, default=2020)
    ap.add_argument("--out", default="enriched_reviewed.json")
    a = ap.parse_args()
    oa.API_KEY = a.api_key

    reviews = read_review_column(a.xlsx, a.sheet, a.name_col, a.inst_col)
    if not reviews:
        sys.exit("「人工核实结果」这一列没有找到任何已填写的内容——确认一下是不是传错文件了，"
                 "或者填完之后忘了保存。")

    data = json.load(open(a.enriched, encoding="utf-8"))
    by_key = {(_norm(r.get("name")), _norm(r.get("institution"))): r for r in data}

    n_id, n_ok, n_skip, n_unrecognized, n_notfound_in_json = 0, 0, 0, 0, 0
    unrecognized_rows = []

    for (name_k, inst_k), raw in reviews.items():
        rec = by_key.get((name_k, inst_k))
        if rec is None:
            n_notfound_in_json += 1
            print(f"⚠️ Excel里「{raw}」这行核实结果，在 {a.enriched} 里找不到对应的人"
                 f"（姓名/机构没对上，可能是Excel被手动改过姓名拼写）：跳过", file=sys.stderr)
            continue

        low = raw.strip().lower()
        try:
            if OPENALEX_ID_RE.match(raw.strip()):
                aid = raw.strip().upper()
                rec.update(build_add(aid, a.mailto, a.window_a_since, "manual_review_confirmed"))
                n_id += 1
                print(f"  ✅ {rec['name']}：确认为 {aid}，已重新抓取完整信息", file=sys.stderr)
            elif low == "ok":
                cands = rec.get("candidates") or []
                if len(cands) != 1:
                    n_unrecognized += 1
                    unrecognized_rows.append((rec["name"], raw,
                                              f"填了'ok'，但这条记录候选人数是{len(cands)}个，不是1个，"
                                              "看不出该确认哪一个——请改填具体的OpenAlex ID"))
                    continue
                aid = cands[0]["id"]
                rec.update(build_add(aid, a.mailto, a.window_a_since,
                                     "manual_review_confirmed_original_candidate"))
                n_ok += 1
                print(f"  ✅ {rec['name']}：确认原候选人 {aid} 正确，已重新抓取完整信息", file=sys.stderr)
            elif low == "skip":
                rec["enrich_status"] = "manual_review_skip"
                rec["match_method"] = "manual_review_skip"
                n_skip += 1
                print(f"  ⏭️ {rec['name']}：确认无法在OpenAlex上找到，标记跳过", file=sys.stderr)
            else:
                n_unrecognized += 1
                unrecognized_rows.append((rec["name"], raw, "格式无法识别"))
        except Exception as e:
            n_unrecognized += 1
            unrecognized_rows.append((rec["name"], raw, f"抓取失败：{type(e).__name__}: {e}"))

    json.dump(data, open(a.out, "w", encoding="utf-8"), ensure_ascii=False, indent=2)

    print(f"\n{'=' * 60}", file=sys.stderr)
    print(f"确认为具体OpenAlex ID：{n_id} 人", file=sys.stderr)
    print(f"确认原候选人正确（填'ok'）：{n_ok} 人", file=sys.stderr)
    print(f"确认跳过（填'skip'）：{n_skip} 人", file=sys.stderr)
    if n_notfound_in_json:
        print(f"⚠️ Excel里填了但在json里找不到对应记录：{n_notfound_in_json} 人", file=sys.stderr)
    if unrecognized_rows:
        print(f"⚠️ 格式无法识别，原样跳过（{len(unrecognized_rows)} 人），请检查这几行的填写格式：",
              file=sys.stderr)
        for name, raw, reason in unrecognized_rows:
            print(f"    - {name}：填的是「{raw}」——{reason}", file=sys.stderr)
    print(f"\nwrote {a.out}", file=sys.stderr)
    print("下一步：把这份文件当成新的 enriched.json，重新跑 confidence.py / "
         "journal_ranking.py / merge_to_excel.py，人工核实过的这些人就会跟自动匹配成功的人"
         "一样，正常参与分档和代表作推荐。", file=sys.stderr)


if __name__ == "__main__":
    main()
