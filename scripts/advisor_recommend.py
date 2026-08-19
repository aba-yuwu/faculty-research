#!/usr/bin/env python3
"""Turn a scored roster into an applicant-facing advisor shortlist.

Three separate outputs, kept apart on purpose (see
references/advisor-recommendation-design.md for the full rationale):

  1. 待人工核实身份 — records this pipeline could not verify are even the
     right person (low identity confidence, contamination risk, or an
     internally-inconsistent research profile). These are pulled OUT of both
     other lists entirely: this pipeline is not confident enough about WHO
     this record is to say anything trustworthy about whether their research
     fits the applicant, so it says nothing rather than guess. Confirming
     identity is a separate manual step (see apply_manual_review.py); once
     confirmed, re-run this script on the corrected data.

  2. 不推荐 — identity is trusted, but the research itself doesn't fit: the
     advisor's recent papers show zero overlap with the applicant's stated
     interests (e.g. applicant wants corporate finance, advisor's last two
     years are agricultural economics), or there is no recent output at all
     to judge fit from.

  3. 推荐排名 — everyone else, ranked by a transparent, reproducible score:
     50% direction fit + 30% journal tier + 20% output intensity. Top-N
     highlighted.

Direction fit (the primary signal, per applicant request) is intentionally
NOT computed as "does the applicant's keyword appear in the paper title" —
title wording varies too much for that to catch genuinely on-topic papers
(a corporate-finance paper titled "Leverage Dynamics and Payout Smoothing in
U.S. Public Firms" contains none of the words "corporate finance"). Instead
it matches against the topic/subfield/field labels OpenAlex's own topic
classifier already assigned each paper (fetch_openalex.py's `topics_detail`)
— a curated taxonomy, not free text — falling back to title-text matching
only for older records that predate that field. This is still keyword/
substring matching, not semantic matching — labeled as such in every reason
string, same convention journal_ranking.py's relevance_score() uses.

Usage:
    python advisor_recommend.py scored.json --interests "corporate finance" \\
        "capital structure" "mergers and acquisitions" --out advisors.xlsx --top-n 10
"""
import argparse
import json
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

# ---------------------------------------------------------------------------
# Direction-fit matching levels. A hit against the paper's specific OpenAlex
# *topic* label is the strongest evidence (same granularity as the interest
# keyword itself); a *subfield* hit is still strong (OpenAlex's own
# sub-discipline bucket, e.g. "Corporate Finance and Governance"); a *field*
# hit is the weakest usable signal (same broad discipline only, e.g. both are
# "Economics, Econometrics and Finance" — not enough alone to call it a fit,
# but not nothing). Values are relative weights, not probabilities.
# ---------------------------------------------------------------------------
LEVEL_SCORE = {"topic": 1.0, "subfield": 0.7, "field": 0.35, "title_fallback": 0.5}

TIER_SCORE = {"top": 100, "good": 70, "default": 45}
TIER_LABEL = {"top": "①高质量期刊作者", "good": "②稳定产出", "default": "④默认档"}

MAX_INTENSITY_PAPERS = 6  # paper count at/above which intensity score saturates at 100


def _match_level(keyword, topic_detail):
    kw = (keyword or "").strip().lower()
    if not kw:
        return None
    topic = (topic_detail.get("topic") or "").lower()
    subfield = (topic_detail.get("subfield") or "").lower()
    field = (topic_detail.get("field") or "").lower()
    if kw in topic:
        return "topic"
    if kw in subfield:
        return "subfield"
    if kw in field:
        return "field"
    return None


def direction_fit(rec, interests):
    """Returns (score 0-100, evidence string). `interests` is required and
    non-empty — direction fit is this module's primary signal, so unlike
    journal_ranking.py's optional --interests, there is no meaningful
    "neutral" fallback when it's missing; the caller enforces this at the CLI
    level rather than silently scoring everyone 0.
    """
    works = rec.get("recent_works") or rec.get("works") or []
    if not works:
        return None, "统计窗口内无任何论文记录，无法判断研究方向是否匹配"

    per_paper, evidence = [], []
    for w in works:
        details = w.get("topics_detail") or []
        best = None
        for kw in interests:
            for d in details:
                lvl = _match_level(kw, d)
                if lvl and (best is None or LEVEL_SCORE[lvl] > LEVEL_SCORE[best[0]]):
                    best = (lvl, kw, d.get("topic"))
        if best:
            lvl, kw, topic_name = best
            per_paper.append(LEVEL_SCORE[lvl])
            evidence.append(
                f"《{w.get('title')}》命中「{kw}」（OpenAlex 主题：{topic_name}，"
                f"匹配层级：{lvl}）"
            )
        elif not details:
            # No topics_detail on this work (older cached record) — fall back
            # to the coarser title-text method journal_ranking.py already
            # uses, rather than treating missing metadata as a mismatch.
            title = (w.get("title") or "").lower()
            hit_kw = next((kw for kw in interests if kw and kw.lower() in title), None)
            if hit_kw:
                per_paper.append(LEVEL_SCORE["title_fallback"])
                evidence.append(f"《{w.get('title')}》标题关键词命中「{hit_kw}」（无主题分类数据，按标题粗匹配）")
            else:
                per_paper.append(0.0)
        else:
            per_paper.append(0.0)

    score = round(100 * sum(per_paper) / len(per_paper), 1)
    if not evidence:
        evidence = [
            f"统计窗口内 {len(works)} 篇论文的 OpenAlex 主题分类与标题均未命中任何一个研究兴趣关键词"
            f"（{', '.join(interests)}）"
        ]
    shown = evidence[:3]
    suffix = f"（共{len(evidence)}条证据，仅展示前3条）" if len(evidence) > 3 else ""
    return score, "；".join(shown) + suffix


def identity_flag_reasons(rec):
    """Reasons this record needs manual identity verification BEFORE its
    research can be judged at all. Kept entirely separate from the fit
    judgment below — a record here says nothing about whether the person is
    a good advisor, only that this pipeline isn't sure yet who the record
    even belongs to.
    """
    reasons = []
    rel = rec.get("reliability") or {}
    level = rel.get("level")
    if level in ("低", "不可用"):
        reasons.append(f"身份可信度{level}（{rel.get('confidence', 0)}分）：{rel.get('action', '需人工核实')}")
    if "profile_contamination_risk" in (rel.get("flags") or []):
        reasons.append("身份记录疑似混入同名他人的成果（跨学科污染风险）")
    jr = rec.get("journal_ranking") or {}
    if jr.get("tier") == "contaminated":
        reasons.append(f"该学者自身近期论文的学科组合内部不自洽（{jr.get('tier_reason', '')}），"
                       f"提示身份可能有误，而非研究方向问题")
    return reasons


def fit_exclusion_reasons(fit_score, paper_count):
    reasons = []
    if paper_count == 0:
        reasons.append("统计窗口内无任何论文产出，无法评估是否适合推荐")
    elif fit_score == 0:
        reasons.append("统计窗口内论文的研究方向与申请者所选兴趣关键词完全没有重合（方向偏移）")
    return reasons


def recommend_score(fit_score, jr, paper_count):
    reasons = [f"研究方向匹配度 {fit_score} 分"]
    tier = (jr or {}).get("tier")
    if tier in TIER_SCORE:
        journal_score = TIER_SCORE[tier]
        reasons.append(f"期刊分级{TIER_LABEL[tier]}：{jr.get('tier_reason', '')}".rstrip("："))
    else:
        journal_score = 50
        reasons.append("未提供 JCR 期刊列表，期刊质量维度按中性分处理")

    n = paper_count
    intensity_score = round(min(100, 100 * n / MAX_INTENSITY_PAPERS), 1)
    reasons.append(f"统计窗口内发表 {n} 篇，密集程度 {intensity_score} 分（{MAX_INTENSITY_PAPERS}篇及以上记满分）")

    score = round(0.5 * fit_score + 0.3 * journal_score + 0.2 * intensity_score, 1)
    return score, "；".join(reasons)


def classify(data, interests, top_n):
    identity_review, excluded, kept = [], [], []
    for rec in data:
        id_reasons = identity_flag_reasons(rec)
        if id_reasons:
            identity_review.append({"rec": rec, "reasons": id_reasons})
            continue

        fit_score, fit_evidence = direction_fit(rec, interests)
        jr = rec.get("journal_ranking") or {}
        n = jr.get("window_a_paper_count")
        n = n if isinstance(n, int) else len(rec.get("recent_works") or rec.get("works") or [])

        excl = fit_exclusion_reasons(fit_score or 0, n)
        if excl:
            excluded.append({"rec": rec, "reasons": excl, "fit_evidence": fit_evidence})
            continue

        score, score_reason = recommend_score(fit_score, jr, n)
        kept.append({
            "rec": rec, "score": score,
            "reason": f"{score_reason}；方向匹配证据：{fit_evidence}",
        })

    kept.sort(key=lambda x: x["score"], reverse=True)
    for i, row in enumerate(kept, 1):
        row["rank"] = i
        row["recommended"] = i <= top_n
    return identity_review, excluded, kept


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
BLOCK_FILL = PatternFill("solid", fgColor="FCE4E4")
REVIEW_FILL = PatternFill("solid", fgColor="FFF2CC")
RECO_FILL = PatternFill("solid", fgColor="E2F0D9")


def _write_sheet(ws, headers, rows, widths, fill=None, highlight_col=None, highlight_fill=None):
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    for row in rows:
        ws.append(row)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    if fill is not None:
        for r in range(2, ws.max_row + 1):
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = fill
    if highlight_col is not None:
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=highlight_col).value:
                for c in range(1, len(headers) + 1):
                    ws.cell(row=r, column=c).fill = highlight_fill


def write_excel(identity_review, excluded, kept, out_path):
    wb = openpyxl.Workbook()

    ws0 = wb.active
    ws0.title = "待人工核实身份"
    rows0 = [
        [e["rec"].get("name"), e["rec"].get("institution"), " / ".join(e["reasons"])]
        for e in identity_review
    ]
    _write_sheet(ws0, ["姓名", "机构", "需人工核实原因"], rows0, [16, 28, 80], fill=REVIEW_FILL)

    ws1 = wb.create_sheet("不推荐")
    rows1 = [
        [e["rec"].get("name"), e["rec"].get("institution"), " / ".join(e["reasons"]), e["fit_evidence"]]
        for e in excluded
    ]
    _write_sheet(ws1, ["姓名", "机构", "不推荐原因", "方向匹配证据"], rows1, [16, 28, 40, 60], fill=BLOCK_FILL)

    ws2 = wb.create_sheet("推荐排名")
    rows2 = [
        [
            k["rank"], "是" if k["recommended"] else "",
            k["rec"].get("name"), k["rec"].get("institution"), k["score"],
            k["reason"],
        ]
        for k in kept
    ]
    _write_sheet(
        ws2, ["排名", "入选Top-N", "姓名", "机构", "综合得分", "推荐理由"],
        rows2, [6, 10, 16, 28, 10, 90],
        highlight_col=2, highlight_fill=RECO_FILL,
    )

    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("scored", help="confidence.py's output, optionally further processed by journal_ranking.py")
    ap.add_argument("--interests", nargs="+", required=True,
                    help="applicant's stated research interests, e.g. --interests \"corporate finance\" \"capital structure\"")
    ap.add_argument("--out", default="advisors.xlsx")
    ap.add_argument("--top-n", type=int, default=10)
    a = ap.parse_args()

    data = json.load(open(a.scored, encoding="utf-8"))
    identity_review, excluded, kept = classify(data, a.interests, a.top_n)
    write_excel(identity_review, excluded, kept, a.out)

    top = [k for k in kept if k["recommended"]]
    print(f"=== 推荐 Top {len(top)} ===", file=sys.stderr)
    for k in top:
        print(f"\n[{k['rank']}] {k['rec'].get('name')}（{k['rec'].get('institution')}）· {k['score']}分", file=sys.stderr)
        print(f"    {k['reason']}", file=sys.stderr)

    print(f"\n=== 不推荐：{len(excluded)} 位 / 待人工核实身份：{len(identity_review)} 位 / 排名池：{len(kept)} 位 ===",
          file=sys.stderr)
    print(f"wrote {a.out}", file=sys.stderr)


if __name__ == "__main__":
    main()
