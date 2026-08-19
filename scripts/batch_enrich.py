#!/usr/bin/env python3
"""Batch-enrich an entire roster from OpenAlex in one run.

Fills the fields a ranking formula actually needs — recent output, venues,
career start, coauthors — for everyone at once, so that ranking is not biased
by how much manual attention each person happened to receive.

Deliberately does NOT try to resolve every ambiguous name. Unresolved rows are
reported for manual handling rather than guessed at.

Input CSV/XLSX must contain at least: id, name, institution
Usage:
  python batch_enrich.py roster.xlsx --sheet Sheet1 --mailto you@x.com --out enriched.json
  python batch_enrich.py roster.json --mailto you@x.com --out enriched.json --window-a-since 2020

Note: --window-a-since replaces the old --since flag. It fetches works from
this (earlier) year onward, which is a superset of journal_ranking.py's
narrower "window B" (its own --window-b-since) — journal_ranking.py re-slices
the same `works` list rather than triggering a second OpenAlex fetch.
"""
import argparse, json, os, sys, time, traceback
from collections import Counter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import fetch_openalex as oa
import resolve_v2 as rv

CACHE = ".openalex_cache.json"


def load_roster(path, sheet=None, cols=None):
    if path.lower().endswith((".xlsx", ".xlsm")):
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[sheet] if sheet else wb.active
        hdr = {str(ws.cell(1, c).value).strip(): c for c in range(1, ws.max_column + 1)
               if ws.cell(1, c).value}
        cmap = cols or {}
        idc = cmap.get("id") or hdr.get("id") or 1
        nmc = cmap.get("name") or hdr.get("name") or 4
        inc = cmap.get("institution") or hdr.get("institution") or 3
        out = []
        for r in range(2, ws.max_row + 1):
            if not ws.cell(r, nmc).value:
                continue
            out.append({"id": ws.cell(r, idc).value,
                        "name": str(ws.cell(r, nmc).value),
                        "institution": str(ws.cell(r, inc).value or "")})
        return out
    return json.load(open(path, encoding="utf-8"))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("roster")
    ap.add_argument("--sheet"); ap.add_argument("--mailto")
    ap.add_argument("--api-key", help="OpenAlex API key (required since Feb 2026 — "
                                      "free at https://openalex.org/settings/api)")
    ap.add_argument("--out", default="enriched.json")
    ap.add_argument("--window-a-since", type=int, dest="since",
                    help="restrict works to this year onward (was --since; renamed because "
                         "journal_ranking.py's decision tree needs this exact cutoff to line "
                         "up with its own WINDOW_A_SINCE)")
    ap.add_argument("--sleep", type=float, default=0.2)
    ap.add_argument("--field", default="finance",
                    help="expected field, used to break ties: finance/accounting/economics/is/om")
    ap.add_argument("--no-cache", action="store_true",
                    help="ignore .openalex_cache.json and re-fetch everyone")
    ap.add_argument("--consecutive-error-threshold", type=int, default=3,
                    help="连续多少人都是api_error才触发熔断冷却，默认3")
    ap.add_argument("--cooldown-seconds", type=float, default=30,
                    help="熔断触发后暂停多少秒，默认30；如果冷却完还是连续失败，"
                        "下一次冷却时间翻倍（最多到--cooldown-max-seconds）")
    ap.add_argument("--cooldown-max-seconds", type=float, default=300,
                    help="冷却时间翻倍的上限，默认300秒(5分钟)")
    a = ap.parse_args()
    oa.API_KEY = rv.API_KEY = a.api_key

    cache = ({} if a.no_cache else
             (json.load(open(CACHE, encoding="utf-8")) if os.path.exists(CACHE) else {}))
    roster = load_roster(a.roster, a.sheet)
    print(f"roster: {len(roster)} rows", file=sys.stderr)

    out, stats = [], Counter()
    t0 = time.time()

    def _print_eta(i):
        # 每20条打印一次"已用时/预计还需多久"，全量跑几百人的时候，光靠
        # 逐条的 [i/N] 状态行不容易估出还要等多久——这个专门给个粗略ETA。
        if i % 20 == 0 or i == len(roster):
            elapsed = time.time() - t0
            rate = elapsed / i if i else 0
            remaining = rate * (len(roster) - i)
            print(f"        …进度 {i}/{len(roster)}，已用时 {elapsed/60:.1f} 分钟，"
                  f"预计还需约 {remaining/60:.1f} 分钟", file=sys.stderr)

    # 跨人熔断状态：resolve_v2.py 里那个"degraded"降级只在同一个人自己的多次
    # 请求之间生效——换到下一个人，又是从头开始整套重试再失败一次，纯粹在
    # 硬撑。如果OpenAlex当时确实不稳定（或者触发了限流），连续好几个人清一色
    # api_error是个明确信号，这时候应该整体暂停一下，而不是继续一个个硬撑重试。
    breaker = {"consecutive_errors": 0, "cooldown": a.cooldown_seconds}

    def _check_circuit_breaker(enrich_status):
        if enrich_status == "api_error" or str(enrich_status).startswith("error"):
            breaker["consecutive_errors"] += 1
        else:
            # 任何一次不是api_error（哪怕是not_found/ambiguous这类正常的"查完了
            # 但没匹配上"），都说明API本身在正常响应，不是系统性故障，重置。
            breaker["consecutive_errors"] = 0
            breaker["cooldown"] = a.cooldown_seconds
            return
        if breaker["consecutive_errors"] >= a.consecutive_error_threshold:
            cd = breaker["cooldown"]
            print(f"\n⏸️ 连续 {breaker['consecutive_errors']} 人都是 api_error，"
                  f"可能是 OpenAlex 暂时不稳定或触发了限流，暂停 {cd:.0f} 秒后继续"
                  f"（不是脚本卡住，是刻意在等）……", file=sys.stderr)
            time.sleep(cd)
            breaker["consecutive_errors"] = 0
            breaker["cooldown"] = min(cd * 2, a.cooldown_max_seconds)

    for i, rec in enumerate(roster, 1):
        # field/since must be part of the key: changing either between runs
        # (e.g. re-running with a different --field, or a different
        # --window-a-since window) must NOT silently reuse a result computed
        # under the old settings just because name+institution match.
        key = f"{rec['name']}|{rec['institution']}|{a.field}|{a.since}"
        if key in cache:
            rec.update(cache[key]); stats["cached"] += 1; out.append(rec)
            _print_eta(i)
            continue

        # note: rv.find_author() already strips honorifics/CJK internally
        # (name_variants), so no separate cleaning is needed here.
        try:
            cands, how = rv.find_author(rec["name"], rec["institution"], a.mailto, field=a.field)
        except Exception as e:
            rec["enrich_status"] = f"error: {e}"
            rec["error_detail"] = traceback.format_exc()[-600:]
            stats["error"] += 1; out.append(rec)
            print(f"  [{i}/{len(roster)}] ERROR      {rec['name']}  -> {type(e).__name__}: {e}",
                  file=sys.stderr)
            print(f"        {rec['error_detail'].strip().splitlines()[-3]}", file=sys.stderr)
            _check_circuit_breaker(rec["enrich_status"])
            _print_eta(i)
            continue

        rec["match_method"] = how
        # "historical_institution_only" means the roster's (current, official-page)
        # institution only shows up in this author's PAST, never as the most recent
        # one on record — i.e. either the person has since moved, or the OpenAlex
        # identity is not clean. "profile_contamination_risk" means this ONE
        # OpenAlex entity's own topics span clearly unrelated fields (e.g. medicine
        # + computer science on the same ID) — a sign OpenAlex's own author
        # disambiguation has absorbed a different real person's work into this ID.
        # Institution/recency checks cannot catch this (it's not a candidate-vs-
        # candidate comparison), so it is checked separately and never auto-accepted
        # even when it is the only candidate found.
        # "field_mismatch_needs_review" means institution+name narrowed to exactly
        # one candidate, but their OpenAlex topics show zero overlap with the
        # roster's configured --field — name+institution matching has a structural
        # blind spot for Chinese names (surname/given-name order isn't consistent
        # across rosters or institutions, so "unique candidate" can still mean
        # "matched the wrong same-two-tokens person" — see pitfalls.md #19), so a
        # single candidate is not treated as automatically trustworthy when an
        # independent signal (field) is available and contradicts it outright.
        # "api_error" means every OpenAlex call for this person raised an exception —
        # this is NOT the same as a genuine zero-result search (see resolve_v2.py's
        # _find_author_inner), and must never collapse into "not_found": a real
        # not_found means "go verify manually, this person may not be well-indexed";
        # api_error means "re-run — this result isn't trustworthy either way."
        is_api_error = how.startswith("api_error")
        is_field_mismatch = how.startswith("field_mismatch_needs_review")
        NEEDS_REVIEW = ("name_only_unverified", "not_found")
        is_historical_only = "_historical_institution_only" in how
        is_contaminated = "_profile_contamination_risk" in how
        if (len(cands) != 1 or how in NEEDS_REVIEW or is_historical_only or is_contaminated
                or is_api_error or is_field_mismatch):
            if is_api_error:
                rec["enrich_status"] = "api_error"
            elif is_field_mismatch:
                rec["enrich_status"] = "needs_review_field_mismatch"
            elif is_contaminated:
                rec["enrich_status"] = "possible_move" if is_historical_only else "needs_review_contaminated"
            elif is_historical_only:
                rec["enrich_status"] = "possible_move"
            else:
                rec["enrich_status"] = "ambiguous" if cands else "not_found"
            rec["candidates"] = [{"id": c["id"], "name": c["name"], "works": c["works"],
                                  "last_known": c["last_known"], "topics": c["topics"],
                                  "contamination_domains": c.get("contamination_domains"),
                                  "url": f"https://openalex.org/{c['id']}"}
                                 for c in cands[:5]]
            import urllib.parse as _u
            rec["verify_url"] = ("https://openalex.org/authors?search="
                                 + _u.quote(str(rec.get("name", ""))))
            stats[rec["enrich_status"]] += 1; out.append(rec)
            print(f"  [{i}/{len(roster)}] {rec['enrich_status'].upper():<10} {rec['name']}"
                  f"  ({how}, {len(cands)} cand)", file=sys.stderr)
            _check_circuit_breaker(rec["enrich_status"])
            _print_eta(i)
            continue

        aid = cands[0]["id"]
        merged_info = {k: cands[0][k] for k in ("merged_ids", "merged_names", "merge_note")
                       if k in cands[0]}
        try:
            prof = oa.profile(aid, a.mailto)
            works = oa.works(aid, a.mailto, a.since)
        except Exception as e:
            rec["enrich_status"] = f"error: {e}"
            rec["error_detail"] = traceback.format_exc()[-600:]
            stats["error"] += 1; out.append(rec)
            print(f"  [{i}/{len(roster)}] ERROR      {rec['name']}  -> {type(e).__name__}: {e}",
                  file=sys.stderr)
            for ln in traceback.format_exc().strip().splitlines()[-3:]:
                print(f"        {ln.strip()}", file=sys.stderr)
            continue

        recent = [w for w in works if (w.get("year") or 0) >= (a.since or 0)]
        add = {
            "openalex_id": aid,
            "match_method": how,
            "orcid": prof.get("orcid"),
            "academic_start_year": prof.get("earliest_affiliation_year"),
            "last_known_institutions": prof.get("last_known_institutions"),
            "effective_institutions": prof.get("effective_institutions"),
            "affiliation_institutions": prof.get("affiliation_institutions"),
            "topics": prof.get("topics"),
            "counts_by_year": prof.get("counts_by_year"),
            "cited_by_count": prof.get("cited_by_count"),          # 历史总被引（非近年窗口）
            "total_works_count": prof.get("works_count"),          # OpenAlex 记录的历史发表论文总数
            "recent_works_count": len(recent),                     # since 年份起的发表数
            "recent_works": [{"title": w["title"], "year": w["year"], "venue": w["venue"],
                              "published": w["is_published"], "doi": w["doi"],
                              "url": w.get("url"), "institutions": w.get("institutions")}
                             for w in recent[:12]],
            "works": works,   # 完整列表（含 url/institutions/coauthors），供导出逐篇明细用
            "enrich_status": "ok",
            **merged_info,
        }
        rec.update(add); cache[key] = add
        stats["ok"] += 1; out.append(rec)
        print(f"  [{i}/{len(roster)}] ok         {rec['name']}  "
              f"({len(recent)} recent works, {how})", file=sys.stderr)
        _check_circuit_breaker(rec["enrich_status"])
        _print_eta(i)
        time.sleep(a.sleep)
        if i % 25 == 0:
            json.dump(cache, open(CACHE, "w", encoding="utf-8"), ensure_ascii=False)
            # 顺手把目前跑出来的部分结果也写一份到 --out：全量跑几百人可能要
            # 一两个小时，中途被打断（Colab断线/手动Ctrl+C）的话，只存缓存不
            # 存输出会导致这几个小时的结果完全拿不到手，还得重新等一遍主循环
            # 走完才有 enriched.json（哪怕缓存命中让这一遍变快，也要重新走一遍
            # 全部行）。这里每25条覆盖写一次，中途断了也有部分结果可以先用。
            json.dump(out, open(a.out, "w", encoding="utf-8"), ensure_ascii=False, indent=2)

    json.dump(cache, open(CACHE, "w", encoding="utf-8"), ensure_ascii=False)
    json.dump(out, open(a.out, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    print(f"\n{dict(stats)}", file=sys.stderr)
    print(f"wrote {a.out}", file=sys.stderr)

    errs = [r["name"] for r in out if r.get("enrich_status") == "api_error"]
    if errs:
        print(f"\n⚠️ {len(errs)} 位教授的OpenAlex请求一直失败（不是查无此人，是网络/接口本身出了问题）："
              f"直接重新运行这一格即可重试——这些人没有写入缓存，会自动重新请求，"
              f"不影响已经跑成功的人：", file=sys.stderr)
        for n in errs[:20]:
            print(f"  - {n}", file=sys.stderr)

    amb = [(r["name"], r.get("enrich_status")) for r in out
          if r.get("enrich_status") not in ("ok", "api_error")]
    if amb:
        print(f"\n{len(amb)} names need manual resolution (not guessed):", file=sys.stderr)
        for n, st in amb[:30]:
            print(f"  - {n}  [{st}]", file=sys.stderr)


if __name__ == "__main__":
    main()
