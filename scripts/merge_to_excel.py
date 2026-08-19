#!/usr/bin/env python3
"""Write enrichment results into a NEW workbook. The original is never modified.

Appends columns rather than overwriting existing ones, so manually curated
assessments survive. Every appended column is prefixed so provenance is obvious.

Also writes a second sheet, "论文明细", with ONE ROW PER PAPER (title, this
author's institution on that specific paper, a link, and coauthors) — a roster
row only has room for a compact preview of recent work, but downstream use
(coauthor network analysis, checking a specific claim) needs the full list.

Usage:
  python merge_to_excel.py original.xlsx enriched.json --sheet 全部教授_论文与聚焦度 \
      --out enriched_output.xlsx
"""
import argparse, json, re, shutil, os, sys
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import resolve_v2 as rv   # reuse the SAME alias table / resolver batch_enrich.py
                          # already used to find these candidates, instead of
                          # reinventing a weaker raw-substring check here.

# NOTE: "近年" here means "since whatever --window-a-since year batch_enrich.py
# was run with" (see recent_works_count / recent_works in the enriched JSON) —
# it is NOT hardcoded to a fixed 3-year window, despite what an earlier
# version of this column's name implied.
#
# 精简过的列表——只保留用户明确要求保留的字段（任教起始年、近年论文清单、研究
# 主题、被引/发表总数、近年被引数、研究方向偏移提示、代表作15列拆分、可信度
# 1-4打分），加上"人工核实结果"这套工作流功能上必需的两列（不然身份核实这个
# 反馈闭环没法运作）。删掉的列（OA_匹配方式、OA_机构时效性、OA_ORCID、OA_当前
# 机构、OA_近年正式发表数、OA_最新发表年、OA_机构是否不一致、OA_合并记录说明、
# 原来的OA_可信度分/等级/建议动作/依据四列、JR_学术分档、JR_分档依据、JR_代表
# 作数、JR_代表作预览）不是丢失了，是不再单独占一列展示——机构一致性检查、
# 可信度的具体加减分依据这些，仍然在内部计算、影响"机构可能不一致"这类汇总
# 统计和最终的1-4打分，只是不再逐行铺开成单独的列。
NEW_COLS = [
    ("OA_匹配状态", "enrich_status"),
    ("OA_候选人(需人工确认)", None),
    ("人工核实结果", None),          # 用户手动填写的核实结论，见 apply_manual_review.py
                                    # 顶部文档字符串里的固定格式说明；生成时永远留空，
                                    # 这一列是给用户填的输入列，不是脚本算出来的输出列
    ("OA_任教起始年", "academic_start_year"),
    ("OA_被引总数(历史总计)", "cited_by_count"),
    ("OA_发表论文总数(历史总计)", "total_works_count"),
    ("OA_近年发表数", None),
    ("OA_近年被引数", None),         # 近年（窗口A起点至今）逐年被引数之和，来自
                                    # OpenAlex的 counts_by_year，不是历史总被引数
    ("OA_近年论文清单(标题/机构/链接见「论文明细」表)", None),
    ("OA_研究主题", None),
    ("JR_研究方向偏移提示", None),   # 简化为固定的几种取值，具体含义见"颜色图例"表
    ("JR_可信度打分(1=最高,4=最低)", None),  # 1=高 2=中 3=低 4=不可用，
                                            # 对应内部可信度评估的四个等级
]
# 代表作最多3篇，按年份从近到远排列；每篇拆成5个独立字段（不是合并成一段文字），
# 一共 3×5=15 列。学科排名用JCR"学科内排名/总数"这个字段（比如"27/266"），
# 不是百分位——用户要求的是"排名"字面意思。
REP_PAPER_FIELDS = ["标题", "合著者", "年份", "期刊名称", "学科排名"]
for _i in range(1, 4):
    for _f in REP_PAPER_FIELDS:
        NEW_COLS.append((f"代表作{_i}_{_f}", None))
del _i, _f

# 0-indexed，对应上面NEW_COLS列表里的位置：1=候选人列表，8=近年论文清单，
# 12/13、17/18、22/23=三篇代表作各自的"标题"和"合著者"（这两个字段内容可能较长，
# 需要换行+加宽；年份/期刊名称/学科排名通常较短，不需要）。
WRAP_COLS = (1, 8, 12, 13, 17, 18, 22, 23)
WIDE_COLS = {1: 45, 8: 60, 12: 32, 13: 32, 17: 32, 18: 32, 22: 32, 23: 32}

PAPER_COLS = ["教授姓名", "名单机构", "论文标题", "发表年份", "期刊/会议",
             "该作者当次所属机构", "文章链接", "DOI", "合作者",
             "JCR分区(ESCI降级后)", "JCR有效百分位", "是否UTD24", "期刊匹配方式", "JCR匹配名"]


def _norm_inst(s):
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


def _institution_mismatch(inst_sheet, last, mailto=None):
    """Does `last` (this author's OpenAlex institution names) actually include
    the roster's institution? Returns "match" | "mismatch" | "unknown".

    A raw substring test on the roster's raw text fails whenever the roster
    uses an acronym ("HKU") that never appears verbatim inside OpenAlex's full
    canonical name ("University of Hong Kong") — that check would flag a
    correct match as a mismatch. Resolve the roster string through the SAME
    alias table / resolver used to find this candidate in the first place
    (resolve_v2.resolve_institution — checks the local alias file before ever
    touching the network, and since batch_enrich.py already resolved every
    institution in this roster, this almost always hits that local cache).

    When resolution genuinely fails (no local alias cached AND the network
    call fails — e.g. OpenAlex API credits exhausted, see pitfalls.md #20),
    falling back to a raw-string substring check is not a safe substitute:
    it can't bridge an abbreviation to its full name any better than the
    check this function exists to replace, so it would silently relabel
    "couldn't verify" as "confirmed mismatch" — a real record of this
    happening flagged 66 people as institution mismatches that were, on
    inspection, all just unresolved abbreviations. "unknown" is reported
    as its own outcome instead, so a genuine mismatch is never diluted by
    (and never confused with) an unrelated API failure.
    """
    if not inst_sheet or not last:
        return "unknown"
    try:
        _, canonical = rv.resolve_institution(inst_sheet, mailto)
    except Exception:
        canonical = None
    if not canonical:
        return "unknown"
    targets = {_norm_inst(canonical), _norm_inst(inst_sheet)}
    targets.discard("")
    for x in last:
        nx = _norm_inst(x)
        if not nx:
            continue
        if nx in targets or any(t and (t in nx or nx in t) for t in targets):
            return "match"
    return "mismatch"


def _recent_citations(rec, since):
    """Sum of per-year citation counts (OpenAlex's counts_by_year — citations
    RECEIVED in that year, not citations to papers published that year) for
    years >= since. Distinct from cited_by_count (historical total) — this is
    specifically "近年被引数", the recent citation trend, not the career total."""
    cby = rec.get("counts_by_year") or {}
    total = 0
    for year, c in cby.items():
        try:
            y = int(year)
        except (TypeError, ValueError):
            continue
        if y >= since:
            total += c.get("cites") or 0
    return total


def _confidence_1_to_4(level):
    """1=最好(高) ... 4=最差(不可用) — a direct relabeling of the existing
    四档(高/中/低/不可用) confidence level, not a new scoring mechanism; see
    confidence.py's score_one() for what actually determines the level."""
    return {"高": 1, "中": 2, "低": 3, "不可用": 4}.get(level)


def _domain_deviation_label(jr):
    """未偏离 / 研究领域较细 / 交叉学科研究 — the three states requested, derived
    from resolve_domain_signal()'s note_detail type (see journal_ranking.py
    §8's rare/bridge branches). A blocked (research-direction-combination-
    doesn't-add-up) record doesn't fit any of the three — representative_papers
    is empty for those anyway, which is itself already a visible signal, so
    this column stays "未偏离" rather than inventing a fourth label the user
    didn't ask for; the block reason itself lives in OA_可信度依据-adjacent
    reasoning inside `reliability`/`journal_ranking`, not lost, just not
    duplicated into this specific column."""
    signal = jr.get("domain_signal") or {}
    note_type = (signal.get("note_detail") or {}).get("type")
    if note_type == "rare":
        return "研究领域较细"
    if note_type == "bridge":
        return "交叉学科研究"
    return "未偏离"


def _rep_paper_columns(jr, max_n=3):
    """3 papers x 5 fields (标题/合著者/年份/期刊名称/学科排名), newest year
    first — flattened into a single list ready to extend() onto `vals`.
    Pads with None if fewer than max_n representative papers exist."""
    reps = sorted(jr.get("representative_papers") or [],
                 key=lambda p: p.get("year") or 0, reverse=True)[:max_n]
    out = []
    for p in reps:
        rank = p.get("jcr_rank")
        rank_total = p.get("jcr_rank_total")
        rank_str = f"{rank}/{rank_total}" if rank and rank_total else ""
        coauthors = "；".join(c.get("name", "") for c in (p.get("coauthors") or []) if c.get("name"))
        out += [p.get("title"), coauthors, p.get("year"), p.get("venue"), rank_str]
    out += [None] * (5 * max_n - len(out))
    return out


def _append_legend_notes(wb):
    """Append explanatory notes for the new/changed columns onto the end of
    the "颜色图例" sheet, if the original workbook has one — this is where a
    reader of the roster would already look for "what does this column mean",
    so new conventions belong there rather than only in a code comment no
    spreadsheet user will ever see."""
    if "颜色图例" not in wb.sheetnames:
        return
    ws = wb["颜色图例"]
    r = ws.max_row + 2
    BOLD = Font(name="微软雅黑", size=9, bold=True)
    BASE = Font(name="微软雅黑", size=9)
    notes = [
        ("补充说明（OA_/JR_/代表作 列）", True),
        ("JR_研究方向偏移提示 —— 固定三种取值：", True),
        ("  未偏离：窗口A论文没有明显偏离主领域，或偏离比例低于阈值", False),
        ("  研究领域较细：偏离比例超过阈值，但机构履历简单，判定为细分方向而非身份问题", False),
        ("  交叉学科研究：偏离的学科跟主领域是常见的桥梁学科组合（如金融+计算机），比例较高但正常", False),
        ("  （研究方向组合本身不成立的极端情况——比如金融+肿瘤医学——代表作会直接是空的，"
         "不会归入以上三种，具体原因见原始 enriched.json 的 journal_ranking.domain_signal 字段）", False),
        ("JR_可信度打分(1=最高,4=最低) —— 对应内部四档可信度评估：", True),
        ("  1 = 高（可直接采用）　2 = 中（建议抽查一眼论文与机构）", False),
        ("  3 = 低（必须人工确认后再用）　4 = 不可用（不要采用，请人工查证）", False),
        ("代表作1/2/3_* —— 最多3篇代表作，按发表年份从近到远排列；每篇拆成5列：", True),
        ("  标题 / 合著者 / 年份 / 期刊名称 / 学科排名（JCR「学科内排名/该学科期刊总数」，"
         "如\"9/243\"表示该学科243本期刊里排第9；不是百分位）", False),
        ("OA_近年被引数 —— 窗口A起点至今，逐年被引数之和（来自OpenAlex的counts_by_year），"
         "跟OA_被引总数(历史总计)是两个不同的统计口径", True),
    ]
    for text, bold in notes:
        c = ws.cell(r, 1)
        c.value = text
        c.font = BOLD if bold else BASE
        c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1


def _write_roster_sheet(wb, ws, data, a):
    norm = lambda s: re.sub(r"\s+", "", str(s or "")).lower()
    # Key on (name, institution), not name alone: a roster with two different
    # people who share a name (common with frequent Chinese surnames) would
    # otherwise have the later one's enrichment silently overwrite the
    # earlier one's in this dict, and BOTH rows would then be merged with
    # whichever one happened to load last.
    by_name_inst = {(norm(r.get("name")), norm(r.get("institution"))): r for r in data}
    # Fallback for the (rare) case a name doesn't reduce to a unique
    # (name, institution) pair here but is otherwise unambiguous in `data`.
    by_name_only = {}
    for r in data:
        k = norm(r.get("name"))
        by_name_only[k] = None if k in by_name_only else r
    start = ws.max_column + 1
    BASE = Font(name="微软雅黑", size=9)
    REVIEW_COL_IDX = [j for j, (t, _) in enumerate(NEW_COLS) if t == "人工核实结果"][0]
    for j, (title, _) in enumerate(NEW_COLS):
        c = ws.cell(1, start + j)
        c.value = title
        c.font = Font(name="微软雅黑", size=9, bold=True)
        if j == REVIEW_COL_IDX:
            # 用不同底色标出这是需要用户填的输入列，不是脚本算出来的输出列
            c.fill = PatternFill("solid", fgColor="FFFFF2CC")
            c.comment = Comment(
                "人工核实完身份之后，在这一列填写结论（留空=还没核实）：\n\n"
                "· 填一个OpenAlex作者ID（如 A5102221172）\n"
                "  = 确认是这个人，会重新抓取这个ID的完整信息\n\n"
                "· 填 ok\n"
                "  = 左边「OA_候选人」列里列出的唯一候选人是对的\n"
                "  （只有候选人正好是1个的时候能这么填）\n\n"
                "· 填 skip\n"
                "  = 确认这个人在OpenAlex上找不到，以后不用再提示\n\n"
                "填完保存，重新上传，运行 apply_manual_review.py 即可。",
                "faculty-research skill")
        else:
            c.fill = PatternFill("solid", fgColor="FFE7F0D9")
    for j, w in WIDE_COLS.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(start + j)].width = w

    n_ok = n_amb = n_moved = n_contaminated = n_field_mismatch = 0
    for r in range(2, ws.max_row + 1):
        nm = ws.cell(r, a.name_col).value
        if not nm:
            continue
        inst_sheet_raw = str(ws.cell(r, a.inst_col).value or "")
        rec = by_name_inst.get((norm(nm), norm(inst_sheet_raw))) or by_name_only.get(norm(nm))
        if not rec:
            continue
        works = rec.get("recent_works") or []
        inst_sheet = inst_sheet_raw
        # effective_institutions already falls back to affiliation history when
        # OpenAlex's own last_known_institutions is empty (which it frequently is,
        # even for authors with a full affiliation record) — using the raw
        # last_known_institutions field alone silently shows "[]" for such rows.
        last = (rec.get("effective_institutions") or rec.get("last_known_institutions")
                or rec.get("affiliation_institutions") or [])
        mismatch = ""
        if rec.get("enrich_status") == "ok" and last and inst_sheet:
            m_state = _institution_mismatch(inst_sheet, last, a.mailto)
            if m_state == "mismatch":
                n_moved += 1
            # "unknown"/"mismatch" no longer get their own displayed column
            # (per the trimmed NEW_COLS) — still counted into n_moved for the
            # summary line, just not repeated into a per-row cell.
        jr = rec.get("journal_ranking") or {}
        rel = rec.get("reliability") or {}
        vals = [
            rec.get("enrich_status"),
            "；".join(f"{c.get('name')}({c.get('works')}篇,{'/'.join(map(str,c.get('last_known') or []))[:28]})"
                     for c in (rec.get("candidates") or [])[:4]),
            "",   # 人工核实结果——留给用户填的输入列，脚本永远不写入内容，但必须
                  # 在这里占一个位置，不然后面所有列的值都会往前错位一格
            rec.get("academic_start_year"),
            rec.get("cited_by_count"),
            rec.get("total_works_count"),
            rec.get("recent_works_count", len(works)),
            _recent_citations(rec, a.window_a_since),
            " || ".join(f"{w.get('year')} {w.get('venue') or '—'} | {str(w.get('title'))[:60]}"
                        f"{' | ' + '/'.join(w.get('institutions') or []) if w.get('institutions') else ''}"
                        for w in works[:8]),
            "；".join((rec.get("topics") or [])[:6]),
            _domain_deviation_label(jr) if jr else "",
            _confidence_1_to_4(rel.get("level")),
        ]
        vals += _rep_paper_columns(jr)
        for j, v in enumerate(vals):
            cell = ws.cell(r, start + j)
            cell.value = v
            cell.font = BASE
            cell.alignment = Alignment(wrap_text=(j in WRAP_COLS), vertical="top")
        if rec.get("enrich_status") == "ok":
            n_ok += 1
        elif rec.get("enrich_status"):
            # Anything that isn't "ok" needs a human look — count it here rather
            # than maintaining a hardcoded status list that drifts out of sync
            # with batch_enrich.py every time a new status is added.
            n_amb += 1
            if rec.get("enrich_status") == "needs_review_contaminated":
                n_contaminated += 1
            if rec.get("enrich_status") == "needs_review_field_mismatch":
                n_field_mismatch += 1
    return n_ok, n_amb, n_moved, n_contaminated, n_field_mismatch


def _write_papers_sheet(wb, data, a):
    """One row per paper across every enriched roster member: title, this
    author's own institution ON that paper, a link, and coauthors — the detail
    a single roster row has no room for.
    """
    ws2 = wb.create_sheet("论文明细")
    BASE = Font(name="微软雅黑", size=9)
    for j, title in enumerate(PAPER_COLS, start=1):
        c = ws2.cell(1, j)
        c.value = title
        c.font = Font(name="微软雅黑", size=9, bold=True)
        c.fill = PatternFill("solid", fgColor="FFE7F0D9")
    for col, width in ((3, 55), (5, 30), (6, 30), (7, 40), (9, 45)):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    r = 2
    n_papers = 0
    for rec in data:
        works = rec.get("works")
        if not works:
            continue
        roster_inst = rec.get("institution") or ""
        for w in works:
            coauthors = "；".join(c.get("name") for c in (w.get("coauthors") or [])
                                 if c.get("name"))
            row = [
                rec.get("name"),
                roster_inst,
                w.get("title"),
                w.get("year"),
                w.get("venue"),
                "；".join(w.get("institutions") or []),
                w.get("url"),
                w.get("doi"),
                coauthors,
                w.get("jcr_quartile_effective"),
                round(w["jcr_percentile"], 1) if isinstance(w.get("jcr_percentile"), (int, float)) else w.get("jcr_percentile"),
                ("是" if w.get("is_utd24") else ("否" if w.get("jcr_matched") is not None else "")),
                w.get("jcr_matched_by"),
                w.get("jcr_matched_name"),
            ]
            for j, v in enumerate(row, start=1):
                cell = ws2.cell(r, j)
                cell.value = v
                cell.font = BASE
                cell.alignment = Alignment(wrap_text=(j in (3, 9)), vertical="top")
            r += 1
            n_papers += 1
    return n_papers


def _write_representative_papers_sheet(wb, data, a):
    """One row per recommended representative paper (up to 3/professor), from
    journal_ranking.py's output. Skipped entirely (sheet not created) if no
    record carries a `journal_ranking` field, so running merge_to_excel.py on
    plain enriched.json (journal ranking step skipped) behaves exactly as before.
    """
    if not any(rec.get("journal_ranking") for rec in data):
        return None
    ws3 = wb.create_sheet("代表作推荐")
    cols = ["教授姓名", "名单机构", "学术分档", "分档依据", "选取依据",
           "论文标题", "发表年份", "期刊/工作论文来源", "是否SSRN工作论文",
           "JCR分区(ESCI降级后)", "JCR有效百分位", "期刊匹配方式", "JCR匹配名", "是否UTD24",
           "排序理由(是否插队)", "相关度说明", "链接", "备注"]
    BASE = Font(name="微软雅黑", size=9)
    for j, title in enumerate(cols, start=1):
        c = ws3.cell(1, j)
        c.value = title
        c.font = Font(name="微软雅黑", size=9, bold=True)
        c.fill = PatternFill("solid", fgColor="FFDCE6F1")
    for col, width in ((4, 45), (5, 45), (6, 50), (13, 40), (15, 35), (16, 30), (18, 35)):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    tier_label = {"top": "①高质量期刊作者", "good": "②成果较好",
                 "not_recommended": "③不建议", "default": "④默认档",
                 "contaminated": "⛔研究方向组合不成立(已跳过推荐)"}
    r = 2
    n_rows = 0
    for rec in data:
        jr = rec.get("journal_ranking")
        if not jr:
            continue
        base = [rec.get("name"), rec.get("institution"),
               tier_label.get(jr.get("tier"), jr.get("tier")),
               jr.get("tier_reason"), jr.get("pick_reason")]
        reps = jr.get("representative_papers") or []
        if not reps:
            # ③不建议 或 找不到任何候选论文时，仍写一行，把判断依据留痕
            row = base + [None] * (len(cols) - len(base))
            for j, v in enumerate(row, start=1):
                cell = ws3.cell(r, j); cell.value = v; cell.font = BASE
                cell.alignment = Alignment(wrap_text=(j in (4, 5)), vertical="top")
            r += 1; n_rows += 1
            continue
        for p in reps:
            row = base + [
                p.get("title"), p.get("year"),
                p.get("venue"), ("是" if p.get("is_ssrn") else "否"),
                p.get("jcr_quartile_effective"), p.get("jcr_percentile_effective"),
                p.get("jcr_matched_by"), p.get("jcr_matched_name"), ("是" if p.get("is_utd24") else "否"),
                p.get("priority_reason"),
                p.get("relevance_note"), p.get("url"), p.get("note"),
            ]
            for j, v in enumerate(row, start=1):
                cell = ws3.cell(r, j)
                cell.value = round(v, 1) if isinstance(v, float) else v
                cell.font = BASE
                cell.alignment = Alignment(wrap_text=(j in (4, 5, 6, 13, 15, 16, 18)), vertical="top")
            r += 1
            n_rows += 1
    return n_rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("original"); ap.add_argument("enriched")
    ap.add_argument("--sheet"); ap.add_argument("--out")
    ap.add_argument("--name-col", type=int, default=4)
    ap.add_argument("--inst-col", type=int, default=3)
    ap.add_argument("--window-a-since", type=int, default=2020,
                    help="算「近年被引数」用的起点年份，跟 journal_ranking.py 的"
                         "--window-a-since 保持一致即可")
    ap.add_argument("--mailto", help="only used if an institution isn't already "
                                     "in the local alias cache from batch_enrich.py")
    ap.add_argument("--api-key", help="OpenAlex API key (required since Feb 2026 — "
                                      "free at https://openalex.org/settings/api)")
    a = ap.parse_args()
    rv.API_KEY = a.api_key

    out = a.out or f"{os.path.splitext(a.original)[0]}_OA补全_{date.today().isoformat()}.xlsx"
    if os.path.abspath(out) == os.path.abspath(a.original):
        raise SystemExit("refusing to overwrite the original file")
    shutil.copy(a.original, out)          # work on a copy; original untouched

    wb = openpyxl.load_workbook(out)
    ws = wb[a.sheet] if a.sheet and a.sheet in wb.sheetnames else wb.active
    data = json.load(open(a.enriched, encoding="utf-8"))

    n_ok, n_amb, n_moved, n_contaminated, n_field_mismatch = _write_roster_sheet(wb, ws, data, a)
    n_papers = _write_papers_sheet(wb, data, a)
    n_reps = _write_representative_papers_sheet(wb, data, a)
    _append_legend_notes(wb)

    wb.save(out)
    print(f"原文件未改动：{a.original}")
    print(f"新文件已生成：{out}")
    print(f"  成功匹配 {n_ok} 行｜待人工确认 {n_amb} 行｜机构可能不一致 {n_moved} 行"
         f"｜疑似 OpenAlex 身份污染 {n_contaminated} 行｜唯一候选人但领域对不上 {n_field_mismatch} 行")
    print(f"  新增 {len(NEW_COLS)} 列（OA_/JR_ 开头的是脚本写入的结果，"
         f"「人工核实结果」是留给你填的输入列，原有列未被覆盖）")
    print(f"  另新增「论文明细」工作表，共 {n_papers} 篇论文（标题/发文机构/链接/合作者）")
    if n_reps is not None:
        print(f"  另新增「代表作推荐」工作表，共 {n_reps} 行（需先跑 journal_ranking.py 才会有此表）")


if __name__ == "__main__":
    main()
