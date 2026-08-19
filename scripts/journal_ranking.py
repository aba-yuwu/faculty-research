#!/usr/bin/env python3
"""Pick up to 3 representative papers per professor, graded by JCR journal quality.

Independent of resolve_v2.py (identity resolution is a separate concern) but
consumes the `works` list batch_enrich.py already fetched — no new OpenAlex
calls are made here.

Decision tree (mutually exclusive, first hit wins — see
journal_ranking_feature_handoff.md §1.1 for the full spec):
  ① Window A has >=1 paper at/above TOP_PERCENTILE_THRESHOLD %ile of its
     category (or hits UTD24) -> tier "top"
       -> pick: window-B SSRN working papers first, else window-B journal
          papers sorted by impact factor
  ② Window A's Q2+ paper ratio > Q2_RATIO_THRESHOLD -> tier "good"
       -> pick: window-B Q2+ journal papers + SSRN, sorted by relevance to
          the user's stated research interests
  ③ Window A's non-SCI ratio > NON_SCI_RATIO_THRESHOLD AND relevance is low
     -> tier "not_recommended" (no papers picked)
  ④ everything else -> tier "default"
       -> pick: window-B journal papers by impact factor desc, padded with
          SSRN working papers if fewer than 3

Usage:
  python journal_ranking.py enriched.json --jcr jcr.xlsx --out ranked.json \
      --window-a-since 2020 --window-b-since 2023 \
      --field-major "Business, Finance" --interests "asset pricing" "ESG"
"""
import argparse, json, os, re, sys
from collections import Counter
from difflib import SequenceMatcher

# ---------------------------------------------------------------------------
# UTD24 — public academic knowledge (not derived from the user's JCR file), so
# per journal_ranking_feature_handoff.md §1.3 it's fine to ship as a constant
# even though the JCR spreadsheet itself must never be bundled or committed.
# Verify against UTD's own page before relying on this for a real cohort —
# the list occasionally gets small revisions.
UTD24_JOURNALS = [
    "The Accounting Review", "Journal of Accounting and Economics",
    "Journal of Accounting Research", "Journal of Finance",
    "Journal of Financial Economics", "Journal of Financial and Quantitative Analysis",
    "Review of Financial Studies", "Information Systems Research", "MIS Quarterly",
    "Journal of Marketing", "Journal of Marketing Research", "Journal of Consumer Research",
    "Journal of Operations Management", "Management Science",
    "Manufacturing & Service Operations Management", "Operations Research",
    "Academy of Management Journal", "Academy of Management Review",
    "Administrative Science Quarterly", "Organization Science",
    "Journal of International Business Studies", "Journal of Applied Psychology",
    "Strategic Management Journal", "Production and Operations Management",
]

_STOPWORDS = {"OF", "AND", "THE", "IN", "FOR", "ON", "A", "AN", "TO"}


def _normalize(name):
    """upper() + strip punctuation + collapse whitespace.

    Shared by BOTH the full-name and the abbreviation match layers (handoff
    §1.7) — comparison only cares about case/punctuation-insensitive equality,
    not the "nice display casing" transform the design doc mentions for
    presentation purposes.
    """
    s = str(name or "").upper()
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _normalize_issn(issn):
    """ISSN format: 4 digits + hyphen + 3 digits + check digit (digit or 'X').
    Normalize by upper-casing and stripping everything but digits/X, so
    '0022-1082' and '0022 1082' and '00221082' all compare equal."""
    s = re.sub(r"[^0-9Xx]", "", str(issn or "")).upper()
    return s


def is_utd24(journal_name):
    if not journal_name:
        return False
    return _normalize(journal_name) in {_normalize(j) for j in UTD24_JOURNALS}


def _to_float(v):
    """Safe float coercion for a raw JCR spreadsheet cell — returns None on
    ANY failure (empty string, non-numeric text, None) rather than raising.
    JCR cells that come through openpyxl as an empty string ('', not None
    and not NaN) are a real, observed data-quality issue — plain float(v)
    crashes the whole run on the first such cell it hits, deep inside
    per-paper annotation, with no indication which journal's row caused it."""
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _parse_category_detail(detail):
    """Parse JCR's '各学科分区详情' cell into per-category records.

    Observed format: one or more entries separated by " | ", each shaped
      "CATEGORY(INDEX_TYPE):QUARTILE RANK/TOTAL"
    e.g. "MANAGEMENT(SSCI):Q1 90/426 | OPERATIONS RESEARCH & MANAGEMENT SCIENCE(SCIE):Q1 19/109"

    Percentile is derived from rank/total using JCR's own formula
    ((total - rank + 1) / total * 100) because the sheet only carries ONE
    top-level 百分位 column even for journals that span multiple categories
    (学科类别 == "Multiple" in that case).
    """
    out = []
    if not detail:
        return out
    for part in str(detail).split("|"):
        part = part.strip()
        m = re.match(r"^(.*?)\(([A-Z]+)\):\s*(Q[1-4])\s+(\d+)/(\d+)$", part)
        if not m:
            continue
        cat, idx_type, quartile, rank, total = m.groups()
        rank, total = int(rank), int(total)
        pct = (total - rank + 1) / total * 100 if total else None
        out.append({"category": cat.strip(), "index_type": idx_type,
                     "quartile": quartile, "rank": rank, "total": total,
                     "percentile": pct})
    return out


def load_jcr(xlsx_path):
    """Parse the user-supplied JCR xlsx into normalized lookup tables.

    Returns {"by_fullname": {...}, "by_abbr": {...}, "by_issn": {...},
    "categories": [...], "count": N}.
    Each record: {name, abbr, issn, eissn, primary_category, jif_percentile,
    jif_quartile, jif, categories: [per-category dicts, see _parse_category_detail]}.

    Per handoff §1.8: this is an in-memory/temp-file-only structure for the
    current run. Callers must never persist the JCR rows themselves (only
    match RESULTS — venue name -> which row matched — may be cached), and
    that cache file must be excluded from version control (see .gitignore).
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(header) if h}

    def g(row, key):
        i = col.get(key)
        return row[i] if i is not None and i < len(row) else None

    by_fullname, by_abbr, by_issn = {}, {}, {}
    categories = set()
    n = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = g(row, "期刊名称")
        if not name:
            continue
        detail = _parse_category_detail(g(row, "各学科分区详情"))
        primary_cat = g(row, "学科类别")
        if primary_cat and str(primary_cat).strip().lower() != "multiple":
            categories.add(str(primary_cat).strip())
        for c in detail:
            categories.add(c["category"])
        rec = {
            "name": name,
            "abbr": g(row, "期刊缩写"),
            "issn": g(row, "ISSN"),
            "eissn": g(row, "eISSN"),
            "primary_category": primary_cat,
            "jif_percentile": _to_float(g(row, "JIF百分位")),
            "jif_quartile": g(row, "JIF分区"),
            "jif": _to_float(g(row, "影响因子JIF")),
            "categories": detail,
        }
        nfull = _normalize(name)
        if nfull:
            by_fullname[nfull] = rec
        nabbr = _normalize(rec["abbr"])
        if nabbr:
            # Verified unique across the full 2026 JCR extract this project
            # ships against (22,643 rows, 0 abbreviation collisions). If a
            # future JCR release ever reuses an abbreviation, the LAST row
            # wins here silently — cheap to re-check on a new file:
            # len(by_abbr) should equal the row count with a non-empty abbr.
            by_abbr[nabbr] = rec
        for issn_field in (rec["issn"], rec["eissn"]):
            nissn = _normalize_issn(issn_field)
            if nissn:
                by_issn[nissn] = rec
        n += 1

    return {"by_fullname": by_fullname, "by_abbr": by_abbr, "by_issn": by_issn,
            "categories": sorted(categories), "count": n}


def match_journal(venue_name, jcr_lookup, venue_issn=None):
    """Layered match: normalized full name -> normalized abbreviation ->
    ISSN -> fuzzy. Returns the JCR record (with an added 'matched_by' key) or
    None.

    venue_issn (optional) is the paper's own ISSN, when the caller has it
    (fetch_openalex.py's works() carries this as `venue_issn`). ISSN is a
    formal identifier, immune to the display-name ambiguity that makes the
    fuzzy layer unsafe to loosen (see references/journal-ranking-design.md —
    e.g. "Journal of Marketing" vs "Journal of Marketing Research" score
    HIGHER on text similarity than a correct but truncated match like
    "Angewandte Chemie" vs its full JCR name, so no text-similarity threshold
    can safely separate those cases; ISSN sidesteps the problem entirely).
    """
    if venue_issn:
        rec = jcr_lookup["by_issn"].get(_normalize_issn(venue_issn))
        if rec:
            return {**rec, "matched_by": "exact_issn"}
    if not venue_name:
        return None
    key = _normalize(venue_name)
    if not key:
        return None
    rec = jcr_lookup["by_fullname"].get(key)
    if rec:
        return {**rec, "matched_by": "exact_fullname"}
    rec = jcr_lookup["by_abbr"].get(key)
    if rec:
        return {**rec, "matched_by": "exact_abbr"}
    # Fuzzy fallback: cheap token-Jaccard prefilter, then SequenceMatcher as
    # the final arbiter (Jaccard alone over-matches short generic names;
    # SequenceMatcher alone can score two DIFFERENT short names deceptively
    # high — e.g. "PLOS ONE" vs "PLOS BIOLOGY" — so both must agree).
    #
    # Deliberately NOT loosened to catch cases like "Angewandte Chemie" vs
    # "ANGEWANDTE CHEMIE-INTERNATIONAL EDITION" (a truncated-but-correct
    # match that scores 0.55 here, below threshold) — tested against real
    # JCR data and confirmed that no jaccard/ratio/token-subset threshold
    # safely separates that from genuinely different journals like "Journal
    # of Marketing" vs "Journal of Marketing Research" (which scores HIGHER,
    # 0.77, despite being the wrong journal). venue_issn above is the correct
    # fix for the truncated-name case; this layer stays conservative.
    best, best_score = None, 0.0
    key_tokens = set(key.split()) - _STOPWORDS
    if key_tokens:
        for cand_key, cand_rec in jcr_lookup["by_fullname"].items():
            cand_tokens = set(cand_key.split()) - _STOPWORDS
            if not cand_tokens:
                continue
            jacc = len(key_tokens & cand_tokens) / len(key_tokens | cand_tokens)
            if jacc < 0.5:
                continue
            ratio = SequenceMatcher(None, key, cand_key).ratio()
            score = (jacc + ratio) / 2
            if score > best_score:
                best, best_score = cand_rec, score
    if best and best_score >= 0.85:
        return {**best, "matched_by": "fuzzy", "fuzzy_score": round(best_score, 3)}
    return None


def _percentile_to_quartile(pct):
    if pct is None:
        return None
    if pct >= 75:
        return "Q1"
    if pct >= 50:
        return "Q2"
    if pct >= 25:
        return "Q3"
    return "Q4"


def effective_percentile(match, field_hint=None):
    """Apply the ESCI -25 percentile penalty (handoff §1.2) and choose which
    category's percentile to use for a (possibly multi-category) journal.

    field_hint, if given, is matched case-insensitively as a substring
    against the journal's per-category list; a hit uses THAT category's own
    percentile/index_type. Otherwise falls back to the category with the
    HIGHEST percentile — the generous default for "Multiple"-category
    journals where the sheet's own 学科类别 column is just the literal
    string "Multiple" and doesn't identify a single field.

    Returns (percentile: float|None, index_type: str|None, category: str|None,
    rank: int|None, total: int|None). rank/total are the JCR "学科内排名"
    (e.g. rank=6, total=140 for "6/140") for the chosen category — this is
    what pick_representative_papers uses to detect a "前3/前5" journal,
    a stronger prestige signal than percentile alone within one category.
    """
    if not match:
        return None, None, None, None, None
    cats = match.get("categories") or []
    chosen = None
    if field_hint:
        fh = field_hint.strip().lower()
        for c in cats:
            if fh in c["category"].lower() or c["category"].lower() in fh:
                chosen = c
                break
    if chosen is None and cats:
        chosen = max(cats, key=lambda c: (c["percentile"] if c["percentile"] is not None else -1))
    if chosen is None:
        # No parseable per-category detail: fall back to the sheet's top-level
        # JIF百分位 (index type unknown -> no ESCI penalty can be applied,
        # i.e. this is the generous case, not the strict one). No rank/total
        # available in this fallback.
        pct = match.get("jif_percentile")
        return pct, None, match.get("primary_category"), None, None
    pct, idx = chosen["percentile"], chosen["index_type"]
    if idx == "ESCI" and pct is not None:
        pct = max(0.0, pct - 25)
    return pct, idx, chosen["category"], chosen.get("rank"), chosen.get("total")


def is_ssrn(work):
    return "ssrn" in str(work.get("venue") or "").lower()


def _paper_text(work):
    return (str(work.get("title") or "") + " " + " ".join(work.get("topics") or [])).lower()


def relevance_score(work, interests):
    """Coarse keyword/phrase overlap between a paper and the user's stated
    research interests — NOT semantic matching (handoff §1.5). Callers must
    label any displayed relevance as keyword-based so it isn't over-trusted.
    """
    if not interests:
        return 0.0
    text_norm = re.sub(r"[^a-z0-9\s]", " ", _paper_text(work))
    hits = 0
    for kw in interests:
        kw_norm = re.sub(r"[^a-z0-9\s]", " ", str(kw or "").lower()).strip()
        if kw_norm and kw_norm in text_norm:
            hits += 1
    return hits / len(interests)


def priority_tier(w):
    """Classify one annotated paper into a priority tier for representative-
    paper selection, plus a human-readable reason for "jumping the queue"
    ahead of a paper with a higher raw percentile.

    Tiers (0 = highest priority):
      0 UTD24        — the handoff spec (§1.3) already says a UTD24 hit
                        "overrides the numeric ranking" for tier classification;
                        applying the same override to *selection order* keeps
                        the two consistent instead of classifying a paper as
                        elite via UTD24 but then ranking it below a merely
                        high-percentile paper when picking representative works.
      1 学科内排名前3  — JCR's own category rank (e.g. "3/109"), a sharper
                        prestige signal than percentile: two journals can both
                        show ~99th percentile while one is #2/2000 and the
                        other #18/2000 in a huge category — rank distinguishes
                        them, percentile alone doesn't.
      2 学科内排名前5  — same idea, slightly wider cutoff.
      3 已匹配JCR      — ordinary case, sorted by percentile within this tier.
      4 未匹配JCR      — always last; never let an unmatched paper outrank a
                        verified one, even if picked as a fallback filler.

    Returns (tier: int, reason: str).
    """
    if w.get("is_utd24"):
        return 0, "UTD24核心期刊（不看具体百分位数值，直接优先）"
    rank, total = w.get("jcr_rank"), w.get("jcr_rank_total")
    if rank is not None:
        if rank <= 3:
            return 1, f"学科类别内排名前3（{rank}/{total}）"
        if rank <= 5:
            return 2, f"学科类别内排名前5（{rank}/{total}）"
    if w.get("jcr_matched"):
        pct = w.get("jcr_percentile")
        return 3, (f"按JCR百分位排序（第{pct:.0f}百分位）" if pct is not None else "已匹配JCR，按百分位排序")
    return 4, "未匹配到JCR，仅作为候选不足时的兜底"


def _annotate(work, jcr_lookup, field_hint):
    """Return a copy of `work` with JCR match info attached."""
    w = dict(work)
    match = match_journal(w.get("venue"), jcr_lookup, venue_issn=w.get("venue_issn"))
    if match is None:
        w.update(jcr_matched=False, jcr_percentile=None, jcr_quartile_effective=None,
                  jcr_index_type=None, jcr_category=None, jcr_matched_by=None, jcr_jif=None,
                  jcr_matched_name=None, jcr_rank=None, jcr_rank_total=None,
                  is_utd24=is_utd24(w.get("venue")))
        w["priority_tier"], w["priority_reason"] = priority_tier(w)
        return w
    pct, idx, cat, rank, total = effective_percentile(match, field_hint)
    w.update(jcr_matched=True, jcr_percentile=pct, jcr_index_type=idx, jcr_category=cat,
              jcr_matched_by=match.get("matched_by"), jcr_jif=match.get("jif"),
              jcr_rank=rank, jcr_rank_total=total,
              # The JCR record's OWN name — may differ from the raw OpenAlex
              # `venue` string above whenever the abbreviation/ISSN/fuzzy
              # layers did the matching (e.g. venue "Angewandte Chemie" but
              # jcr_matched_name "ANGEWANDTE CHEMIE-INTERNATIONAL EDITION").
              # This is the string to search for IN THE JCR SPREADSHEET —
              # searching for the raw venue name instead is exactly what
              # makes an ISSN/abbr/fuzzy-matched paper look "not in the file."
              jcr_matched_name=match.get("name"),
              is_utd24=(is_utd24(w.get("venue")) or is_utd24(match.get("name"))),
              jcr_quartile_effective=_percentile_to_quartile(pct))
    w["priority_tier"], w["priority_reason"] = priority_tier(w)
    return w


def annotate_works(works, jcr_lookup, field_hint):
    """Annotate EVERY work with JCR match info, regardless of window.

    Call this once per professor and reuse the result everywhere downstream
    (classify_professor, pick_representative_papers, and — importantly — the
    full per-paper list persisted back into `rec["works"]` in main(), which is
    what merge_to_excel.py's "论文明细" sheet reads). Previously each paper
    was only annotated transiently inside classify_professor/pick_representative_
    papers and the annotation was thrown away afterwards, so nothing outside
    the 3 picked "representative papers" ever showed a JCR quartile — the
    small-sample preview cell in the notebook (which reads batch_enrich.py's
    raw, chronological `recent_works` and never touches this module at all)
    could look like an unfiltered "recommendation," which it never was.
    """
    return [_annotate(w, jcr_lookup, field_hint) for w in works]


def _evidence(w):
    """Compact, auditable snapshot of one paper for tier_detail — lets the
    user check the ACTUAL JCR record a paper matched to (which can have a
    different name than the raw OpenAlex venue string, e.g. via the
    abbreviation/ISSN/fuzzy layers), not just a bare title they'd have to go
    hunt for in the works list themselves.
    """
    return {"title": w.get("title"), "year": w.get("year"), "venue": w.get("venue"),
           "jcr_matched_name": w.get("jcr_matched_name"), "matched_by": w.get("jcr_matched_by"),
           "percentile": w.get("jcr_percentile"), "quartile": w.get("jcr_quartile_effective"),
           "rank": w.get("jcr_rank"), "rank_total": w.get("jcr_rank_total"),
           "priority_reason": w.get("priority_reason"),
           "is_utd24": w.get("is_utd24")}


# ---------------------------------------------------------------------------
# 17 domain buckets, grouped into 5 families for the relationship table below
# (bucket_relationship). Deliberately keeps 商科/经济/管理 as one coarse bucket
# (finance/econ/management/marketing/accounting/OR are treated as one
# internally-coherent field for this purpose) while splitting the other
# families into their conventional sub-disciplines — see
# references/journal-ranking-design.md §8 for the full rationale and the
# specific example cases (neuroeconomics, agricultural economics, etc.) that
# drove each split and exception.
#
# A category is assigned to the FIRST bucket whose keyword matches; more
# specific terms are listed first where two buckets could otherwise both
# plausibly match (e.g. "COMPUTER SCIENCE" is checked in a bucket ordered
# before "ENGINEERING", so "COMPUTER SCIENCE, SOFTWARE ENGINEERING" lands in
# 计算机科学, not 工程 — though _domain_bucket's prefix-first strategy, see
# below, already resolves most of these before bucket order even matters).
DOMAIN_KEYWORDS = [
    ("商科/经济/管理", ["BUSINESS", "ECONOM", "FINANC", "MANAGEMENT", "MARKETING",
                     "ACCOUNT", "OPERATIONS RESEARCH", "INDUSTRIAL RELATIONS"]),
    ("神经科学", ["NEUROSCIENC", "NEUROLOG"]),
    ("医学/临床", ["MEDICIN", "MEDICAL", "CLINICAL", "ONCOLOG", "CANCER", "CARDIOLOG", "SURGERY",
                "IMMUNOLOG", "PATHOLOG", "RADIOLOG", "PEDIATRIC", "PSYCHIATRY",
                "DERMATOLOG", "UROLOG", "GASTROENTEROLOG", "HEMATOLOG",
                "ENDOCRINOLOG", "NEPHROLOG", "OBSTETRIC", "OPHTHALMOLOG",
                "OTORHINOLARYNGOLOG", "ANESTHESIOLOG", "DENTISTRY", "GERIATRIC",
                "ALLERGY", "ANDROLOGY", "VIROLOG", "TROPICAL MEDICINE", "NURSING"]),
    ("化学", ["CHEMISTRY", "ELECTROCHEMISTRY", "PHARMACOLOG", "TOXICOLOG"]),
    ("生物学/生命科学", ["BIOLOG", "BIOCHEMISTRY", "GENETIC", "MICROBIOLOG",
                     "TISSUE ENGINEERING", "BIOTECHNOLOG", "ZOOLOGY"]),
    ("农业/食品科学", ["PLANT SCIENCE", "VETERINARY", "AGRICULTUR", "FOOD SCIENCE",
                   "NUTRITION", "ENTOMOLOGY"]),
    ("物理/天文", ["PHYSICS", "ASTRONOMY", "NUCLEAR", "OPTICS", "ACOUSTICS"]),
    ("材料科学", ["MATERIALS SCIENCE", "METALLURG"]),
    ("计算机科学", ["COMPUTER SCIENCE"]),
    ("数学/统计", ["MATHEMATIC", "STATISTICS"]),
    ("工程", ["ENGINEERING", "TELECOMMUNICATIONS", "ROBOTICS", "AUTOMATION"]),
    ("心理学", ["PSYCHOLOG"]),
    ("社会学/政治学/传播学", ["SOCIOLOG", "POLITICAL SCIENCE", "ANTHROPOLOG", "DEMOGRAPHY",
                        "CRIMINOLOG", "LINGUISTIC", "COMMUNICATION", "SOCIAL WORK",
                        "URBAN STUDIES"]),
    ("法学", ["LAW"]),
    ("教育学", ["EDUCATION", "LIBRARY SCIENCE"]),
    ("人文", ["HISTORY", "PHILOSOPHY", "ETHICS", "RELIGION"]),
    ("地球/环境科学", ["GEOLOG", "GEOCHEMISTRY", "GEOPHYSICS", "GEOGRAPHY", "ENVIRONMENTAL",
                   "METEOROLOG", "OCEANOGRAPHY", "ECOLOGY", "FORESTRY", "MINING",
                   "ENERGY & FUELS", "WATER RESOURCES"]),
]


def _domain_bucket(category):
    """Map a JCR category string to one of the 17 domain buckets, or None if
    no keyword matches — e.g. "MULTIDISCIPLINARY SCIENCES" is deliberately
    left unbucketed rather than guessed, since a paper in a multidisciplinary
    journal says nothing about which field the AUTHOR is actually in.

    JCR categories consistently follow a "MAIN FIELD, SUBFIELD" naming
    convention (e.g. "PHYSICS, MATHEMATICAL", "COMPUTER SCIENCE, SOFTWARE
    ENGINEERING"). The part before the first comma is checked FIRST, across
    all buckets, before falling back to the full string — this is what
    correctly sends "PHYSICS, MATHEMATICAL" to 物理/天文 rather than
    数学/统计, without needing to hand-tune bucket check order.

    Matches also require a word boundary immediately BEFORE the keyword (but
    not necessarily after, since several keywords are deliberate prefixes
    meant to catch multiple word forms — e.g. "ECONOM" catches both
    "ECONOMICS" and "ECONOMIST"). Plain substring matching without this
    check is a real bug caught by auditing the full 254-category list (see
    pitfalls.md #15): "CHEMISTRY" was matching inside "GEOCHEMISTRY &
    GEOPHYSICS" — an earth-science category with no boundary before
    "CHEMISTRY" — wrongly bucketing it as chemistry.
    """
    if not category:
        return None
    cat = category.upper()
    prefix = cat.split(",", 1)[0].strip()
    for probe in ((prefix, cat) if prefix != cat else (cat,)):
        for bucket, keywords in DOMAIN_KEYWORDS:
            if any(re.search(rf"(?:^|[^A-Z]){re.escape(kw)}", probe) for kw in keywords):
                return bucket
    return None


# ---------------------------------------------------------------------------
# Pairwise relationship between two domain buckets — "reject" (essentially
# impossible for one real scholar), "rare" (a real but small crossover
# field), or "bridge" (a common, unremarkable combination). Looked up in two
# steps: (1) a small table of specific EXCEPTIONS for named pairs where the
# family-level default below would be wrong, checked first; (2) each
# bucket's FAMILY_OF default relationship, used for every pair not called
# out as an exception. This keeps the maintenance burden at "5 families x 5
# families" plus a short, auditable exception list, instead of a full
# 17x17 = 136-pair hand-built table that would need updating every time a
# bucket gets added or split further.
FAMILY_OF = {
    "商科/经济/管理": "A",
    "医学/临床": "B", "神经科学": "B", "化学": "B", "生物学/生命科学": "B",
    "农业/食品科学": "B", "物理/天文": "B", "材料科学": "B",
    "计算机科学": "C", "数学/统计": "C", "工程": "C",
    "心理学": "D", "社会学/政治学/传播学": "D", "法学": "D", "教育学": "D", "人文": "D",
    "地球/环境科学": "E",
}

# A=商科, B=硬科学, C=量化, D=人文社科, E=地球/环境科学. Same-family pairs
# (frozenset of one letter, e.g. {"B"}) default to "bridge" — shared academic
# training/methodology within a family makes co-occurrence unremarkable.
_FAMILY_DEFAULT = {
    frozenset("AB"): "reject", frozenset("AC"): "bridge",
    frozenset("AD"): "rare",   frozenset("AE"): "bridge",
    frozenset("B"):  "bridge",
    frozenset("BC"): "bridge", frozenset("BD"): "rare", frozenset("BE"): "bridge",
    frozenset("C"):  "bridge",
    frozenset("CD"): "bridge", frozenset("CE"): "bridge",
    frozenset("D"):  "bridge",
    frozenset("DE"): "rare",
}

# Specific pairs where the family default above is wrong for THIS pair,
# because a real, named, established interdisciplinary field connects them
# (overriding a stricter default) or, conversely, no such field connects them
# despite a more lenient family default. Each entry is commented with the
# reasoning so a future edit knows whether it's still warranted.
RELATIONSHIP_EXCEPTIONS = {
    # A-B overrides: named crossover fields that the blanket A-B "reject"
    # default would otherwise miss.
    frozenset({"商科/经济/管理", "神经科学"}): "bridge",         # neuroeconomics
    frozenset({"商科/经济/管理", "农业/食品科学"}): "rare",       # agricultural economics
    frozenset({"商科/经济/管理", "材料科学"}): "rare",           # critical-minerals / supply-chain economics
    # B-D overrides: two real, large crossover fields the blanket B-D "rare"
    # default would undersell.
    frozenset({"神经科学", "心理学"}): "bridge",                # cognitive neuroscience
    frozenset({"医学/临床", "心理学"}): "bridge",                # health / clinical psychology
    # B-B overrides: two small, real crossover fields that don't rise to the
    # same-family "bridge" default's level of "unremarkable."
    frozenset({"农业/食品科学", "物理/天文"}): "rare",           # agricultural/soil physics
    frozenset({"农业/食品科学", "材料科学"}): "rare",            # agri-biomaterials
    # 物理/天文 and 材料科学 vs the D family: no established crossover field
    # connects hard physical science with any of these — stricter than the
    # blanket B-D "rare" default.
    frozenset({"物理/天文", "心理学"}): "reject",
    frozenset({"物理/天文", "社会学/政治学/传播学"}): "reject",
    frozenset({"物理/天文", "法学"}): "reject",
    frozenset({"物理/天文", "教育学"}): "reject",
    frozenset({"物理/天文", "人文"}): "reject",
    frozenset({"材料科学", "心理学"}): "reject",
    frozenset({"材料科学", "社会学/政治学/传播学"}): "reject",
    frozenset({"材料科学", "法学"}): "reject",
    frozenset({"材料科学", "教育学"}): "reject",
    frozenset({"材料科学", "人文"}): "reject",
}


def bucket_relationship(bucket_a, bucket_b):
    """"reject" | "rare" | "bridge" for a pair of domain buckets — see
    RELATIONSHIP_EXCEPTIONS / _FAMILY_DEFAULT above for how this is derived."""
    if bucket_a == bucket_b:
        return "bridge"
    key = frozenset({bucket_a, bucket_b})
    if key in RELATIONSHIP_EXCEPTIONS:
        return RELATIONSHIP_EXCEPTIONS[key]
    return _FAMILY_DEFAULT[frozenset({FAMILY_OF[bucket_a], FAMILY_OF[bucket_b]})]


MIN_TOTAL_FOR_RATIO_CHECK = 4
# Below this many JCR-matched window-A papers (after any single-paper reject
# exclusion, see resolve_domain_signal), the "rare"/"bridge" ratio checks are
# skipped — with too few papers, a ratio isn't a meaningful statement about
# this professor's actual output mix. The "reject" branch is NOT gated by
# this floor: it's a count-based rule (>1 paper), not a ratio, so it doesn't
# need a sample-size floor to be meaningful.

RARE_RATIO_THRESHOLD = 0.30
BRIDGE_RATIO_THRESHOLD = 0.50
LOW_INSTITUTION_COUNT_MAX = 4
# At/below this many career institutions (rec["affiliation_institutions"],
# OpenAlex's full affiliation history — not just current employer), a
# rare/bridge ratio above threshold is read as "this professor's career has
# a legible secondary focus" rather than a red flag; above it, the same
# ratio is read as inconsistent enough with a single coherent career to
# withhold recommendations pending manual review.


def resolve_domain_signal(works_a_annotated, institution_count=None, field_hint=None):
    """The single entry point for the research-direction check. Decides,
    for one professor's window-A journal papers:
      - whether any bucket should be EXCLUDED from downstream tier/recommendation
        processing entirely (a lone "reject"-relationship paper — treated as
        JCR misclassification noise, see pitfalls.md #15/#17);
      - whether the professor's whole record should be BLOCKED from tier
        classification and representative-paper recommendation (>1 "reject"
        papers, or a "rare"/"bridge" ratio high enough combined with a
        career spanning more than LOW_INSTITUTION_COUNT_MAX institutions);
      - or, short of either, an informational NOTE to attach alongside
        otherwise-normal tier/recommendation processing.

    Must be called BEFORE classify_professor()/pick_representative_papers()
    — its "excluded_buckets" result has to be applied to works_a AND works_b
    before either of those run, not after (see main()). This is why the
    domain-mix logic no longer lives inside classify_professor: exclusion
    changes what "this professor's papers" even means for every downstream
    step, so it has to be resolved first, once, not threaded through two
    separate functions' decision trees.

    Returns {"primary": bucket|None, "excluded_buckets": [...], "block": bool,
    "block_reason": str|None, "note": str|None, "note_detail": {...}|None}.
    """
    by_bucket = {}
    for w in works_a_annotated:
        b = _domain_bucket(w.get("jcr_category"))
        if b:
            by_bucket.setdefault(b, []).append(w)

    result = {"primary": None, "excluded_buckets": [], "block": False,
             "block_reason": None, "note": None, "note_detail": None}
    if len(by_bucket) < 2:
        return result

    primary = None
    if field_hint:
        hint_bucket = _domain_bucket(field_hint)
        if hint_bucket in by_bucket:
            primary = hint_bucket
    if primary is None:
        primary = max(by_bucket, key=lambda b: len(by_bucket[b]))
    result["primary"] = primary

    reject_buckets, rare_buckets, bridge_buckets = {}, {}, {}
    for b, ws in by_bucket.items():
        if b == primary:
            continue
        rel = bucket_relationship(primary, b)
        {"reject": reject_buckets, "rare": rare_buckets, "bridge": bridge_buckets}[rel][b] = ws

    # ---- ❌ reject branch: count-based, no minimum-sample floor ----
    reject_papers = [w for ws in reject_buckets.values() for w in ws]
    if len(reject_papers) > 1:
        examples = "；".join(f"《{str(w.get('title'))[:30]}》({w.get('venue')})" for w in reject_papers[:4])
        result["block"] = True
        result["block_reason"] = (
            f"窗口A论文中，「{primary}」与「{' / '.join(sorted(reject_buckets))}」这几个学科基本不"
            f"可能同时是同一学者的真实研究方向（共{len(reject_papers)}篇：{examples}），研究方向"
            "组合本身不成立，已跳过分档与代表作推荐，建议人工核实身份。")
        return result
    if len(reject_papers) == 1:
        excluded = next(iter(reject_buckets))
        result["excluded_buckets"] = [excluded]
        w = reject_papers[0]
        result["note"] = (
            f"已自动排除「{excluded}」下的1篇论文（《{str(w.get('title'))[:30]}》，{w.get('venue')}）"
            f"——与主领域「{primary}」基本不可能是同一学者的研究方向组合，更可能是期刊学科归类"
            "噪音，不参与分档与代表作推荐。")
        result["note_detail"] = {"type": "excluded", "bucket": excluded,
                                 "title": w.get("title"), "venue": w.get("venue")}
        # fall through — still check rare/bridge on what's left

    total = sum(len(ws) for ws in by_bucket.values()) - len(reject_papers)
    if total < MIN_TOTAL_FOR_RATIO_CHECK:
        return result

    def _ratio_outcome(label, papers, threshold, low_inst_message):
        ratio = len(papers) / total
        if ratio <= threshold:
            return None
        if institution_count is not None and institution_count <= LOW_INSTITUTION_COUNT_MAX:
            return ("note", f"{low_inst_message}（{label}占比{ratio:.0%}）")
        return ("block",
               f"窗口A论文中「{label}」占比达到{ratio:.0%}，且作者对应机构数超过"
               f"{LOW_INSTITUTION_COUNT_MAX}个，研究方向差异较大，已跳过分档与代表作推荐，"
               "建议人工核实身份。")

    # priority: only the first branch that fires anything (note OR block) is
    # reported — 🟡rare is checked before ✅bridge, and if 🟡rare fires at
    # all (even just a "note", not a block), ✅bridge is not checked.
    rare_papers = [w for ws in rare_buckets.values() for w in ws]
    if rare_papers:
        outcome = _ratio_outcome(" / ".join(sorted(rare_buckets)), rare_papers,
                                 RARE_RATIO_THRESHOLD, "研究领域较细、针对性较高")
        if outcome:
            kind, text = outcome
            if kind == "block":
                result["block"], result["block_reason"] = True, text
                return result
            result["note"] = text
            result["note_detail"] = {"type": "rare", "domains": sorted(rare_buckets),
                                     "ratio": round(len(rare_papers) / total, 3)}
            return result

    bridge_papers = [w for ws in bridge_buckets.values() for w in ws]
    if bridge_papers:
        outcome = _ratio_outcome(" / ".join(sorted(bridge_buckets)), bridge_papers,
                                 BRIDGE_RATIO_THRESHOLD, "该作者倾向交叉学科研究")
        if outcome:
            kind, text = outcome
            if kind == "block":
                result["block"], result["block_reason"] = True, text
                return result
            result["note"] = text
            result["note_detail"] = {"type": "bridge", "domains": sorted(bridge_buckets),
                                     "ratio": round(len(bridge_papers) / total, 3)}

    return result


def classify_professor(works_a_annotated, thresholds, interests=None):
    """Run the ①②③④ decision tree over a professor's window-A journal papers.

    works_a_annotated must already be: (1) filtered to window A years,
    published journal articles (SSRN/working papers excluded — the tree only
    looks at actual journal output for the quality judgment, see handoff
    §1.1), (2) already passed through annotate_works(), and (3) already had
    any resolve_domain_signal()-excluded buckets filtered out — see main().
    This function has no domain-mix awareness of its own; that check runs
    once, earlier, in main(), because its "exclude this bucket" outcome has
    to change what "this professor's papers" means for BOTH tier
    classification and representative-paper picking, not just one of them.

    Returns (tier, detail) where tier in {"top","good","not_recommended","default"}
    and detail always carries a human-readable `reason` string.
    """
    top_pct_th = thresholds.get("top_percentile", 90)
    q2_ratio_th = thresholds.get("q2_ratio", 0.5)
    non_sci_ratio_th = thresholds.get("non_sci_ratio", 0.25)

    annotated = works_a_annotated
    n = len(annotated)

    if n == 0:
        return "default", {"reason": "窗口A内没有期刊论文记录，无法判断成果质量，按默认档处理",
                           "n_window_a": 0}

    # ① 高质量期刊作者
    top_hits = [w for w in annotated
                if w["is_utd24"] or (w["jcr_percentile"] is not None and w["jcr_percentile"] >= top_pct_th)]
    if top_hits:
        utd24_hits = [w for w in top_hits if w["is_utd24"]]
        # Cite an example paper that actually satisfies whichever condition is
        # named as "via" — a paper can land in top_hits through percentile
        # alone, so top_hits[0] is not necessarily a UTD24 hit even when
        # OTHER papers in the list are; citing the wrong one here would make
        # this reason string misleading exactly where it needs to be auditable.
        if utd24_hits:
            via, example = "UTD24", utd24_hits[0]
        else:
            via, example = f"JCR百分位≥{top_pct_th}", top_hits[0]
        return "top", {
            "reason": f"窗口A内存在命中「{via}」的论文（如《{str(example.get('title') or '')[:40]}》），判定为高质量期刊作者",
            "n_window_a": n, "evidence": [_evidence(w) for w in top_hits[:5]]}

    # ② 成果较好：effective percentile>=50 视为 Q2+
    q2plus = [w for w in annotated if w["jcr_percentile"] is not None and w["jcr_percentile"] >= 50]
    q2_ratio = len(q2plus) / n
    if q2_ratio > q2_ratio_th:
        return "good", {"reason": f"窗口A内Q2+期刊论文占比 {q2_ratio:.0%}（阈值 {q2_ratio_th:.0%}），判定为成果较好",
                        "n_window_a": n, "q2_ratio": round(q2_ratio, 3),
                        "evidence": [_evidence(w) for w in q2plus[:5]]}

    # ③ 不建议：非SCI占比超阈值 且 (已知)相关度低。相关度无法判断（未填研究方向）时不触发③，退到④。
    non_sci = [w for w in annotated if not w["jcr_matched"] or w["jcr_index_type"] == "ESCI"]
    non_sci_ratio = len(non_sci) / n
    if interests:
        avg_rel = sum(relevance_score(w, interests) for w in annotated) / n
        if non_sci_ratio > non_sci_ratio_th and avg_rel == 0:
            return "not_recommended", {
                "reason": f"窗口A内非SCI/ESCI/未匹配论文占比 {non_sci_ratio:.0%}（阈值 {non_sci_ratio_th:.0%}）"
                          f"，且论文主题与所填研究方向关键词均无匹配",
                "n_window_a": n, "non_sci_ratio": round(non_sci_ratio, 3), "avg_relevance": avg_rel,
                "evidence": [_evidence(w) for w in non_sci[:5]]}

    return "default", {"reason": "未命中①②③任一条件，按默认档处理", "n_window_a": n,
                       "q2_ratio": round(q2_ratio, 3), "non_sci_ratio": round(non_sci_ratio, 3)}


def pick_representative_papers(works_b_annotated, tier, interests, max_n=3):
    """Pick up to `max_n` representative papers from window-B works, following
    the recommendation rule tied to `tier` (handoff §1.1).

    works_b_annotated must already be annotate_works()'d (same rationale as
    classify_professor above — annotate once, reuse everywhere).

    Returns (picked: list[dict annotated works], pick_reason: str).
    """
    ssrn, journal = [], []
    for w in works_b_annotated:
        if is_ssrn(w):
            w = dict(w, ssrn_note="SSRN工作论文——仅覆盖已被OpenAlex索引的条目，只挂在作者个人主页、"
                                  "尚未被索引的工作论文抓不到")
            ssrn.append(w)
        else:
            journal.append(w)

    def by_jcr_rank(ws):
        # Sort by (priority_tier ascending, percentile descending) instead of
        # raw impact factor: IF is not comparable across fields (a finance
        # journal's IF~5 vs a medicine journal's IF~50 says nothing about
        # relative prestige), so sorting representative-paper candidates by
        # raw jcr_jif silently favoured high-IF fields regardless of subject.
        # priority_tier already encodes the "does this paper jump the queue"
        # logic (UTD24 / 学科前3 / 学科前5 override plain percentile ranking —
        # see priority_tier()'s docstring for why each override exists), so
        # sorting on it directly makes that override visible in the output
        # order, not just implicit in a numeric comparison.
        return sorted(ws, key=lambda w: (w["priority_tier"], -(w["jcr_percentile"] or -1)))

    def by_relevance(ws):
        return sorted(ws, key=lambda w: relevance_score(w, interests), reverse=True)

    if tier == "top":
        picked = ssrn[:max_n]
        if len(picked) < max_n:
            picked = picked + by_jcr_rank(journal)[:max_n - len(picked)]
        reason = "①高质量期刊作者：优先展示窗口B内的SSRN工作论文，不足则按JCR排名从窗口B期刊论文补齐"
    elif tier == "good":
        q2plus_journal = [w for w in journal
                          if w.get("jcr_percentile") is not None and w["jcr_percentile"] >= 50]
        picked = by_relevance(q2plus_journal + ssrn)[:max_n]
        reason = "②成果较好：从窗口B的Q2+期刊论文与SSRN工作论文中，按与研究方向的关键词相关度排序"
    elif tier == "not_recommended":
        picked, reason = [], "③不建议：按判断依据不推荐该教授，未生成代表作列表"
    else:
        picked = by_jcr_rank(journal)[:max_n]
        if len(picked) < max_n:
            picked = picked + ssrn[:max_n - len(picked)]
        reason = "④默认档：窗口B期刊论文按JCR排名从高到低取前3篇，不足则用SSRN工作论文补齐"

    for w in picked:
        w["relevance_note"] = ("相关度：基于关键词匹配，仅供参考" if interests
                               else "未填写研究方向关键词，未计算相关度")
    return picked, reason


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("enriched", help="batch_enrich.py 的输出 json")
    ap.add_argument("--jcr", required=True, help="用户上传的 JCR 期刊名单 xlsx")
    ap.add_argument("--out", default="ranked.json")
    ap.add_argument("--window-a-since", type=int, default=2020,
                    help="判断'是否有高质量期刊'的窗口起点")
    ap.add_argument("--window-b-since", type=int, default=2023,
                    help="真正候选论文/工作论文的窗口起点")
    ap.add_argument("--top-percentile", type=int, default=90,
                    help="①档：前多少百分位算高质量，默认90（前10%）")
    ap.add_argument("--q2-ratio", type=float, default=0.5,
                    help="②档：Q2+占比阈值，默认0.5")
    ap.add_argument("--non-sci-ratio", type=float, default=0.25,
                    help="③档：非SCI占比阈值，默认0.25")
    ap.add_argument("--field-major", default=None,
                    help="学科类别大类，用于多学科('Multiple')期刊选取百分位口径，"
                         "也用于研究方向信号里锚定'主领域'")
    ap.add_argument("--interests", nargs="*", default=None,
                    help="1-3个研究方向关键词/短语")
    a = ap.parse_args()

    if a.window_b_since < a.window_a_since:
        sys.exit(f"配置错误：--window-b-since ({a.window_b_since}) 不能早于 "
                 f"--window-a-since ({a.window_a_since})。窗口B应该是窗口A的一个更晚、"
                 "更窄的子区间（用来挑真正的代表作），窗口A应该是更早、更宽的区间"
                 "（用来判断整体成果质量）——配反了会导致窗口B里出现窗口A没统计到的论文，"
                 "让研究方向信号的排除逻辑失效。请检查配置后重跑。")

    print(f"载入JCR数据：{a.jcr} ...", file=sys.stderr)
    jcr_lookup = load_jcr(a.jcr)
    print(f"  共 {jcr_lookup['count']} 条期刊记录，{len(jcr_lookup['categories'])} 个学科类别",
          file=sys.stderr)

    data = json.load(open(a.enriched, encoding="utf-8"))
    thresholds = {"top_percentile": a.top_percentile, "q2_ratio": a.q2_ratio,
                  "non_sci_ratio": a.non_sci_ratio}

    match_stats = Counter()
    n_scored = 0
    for rec in data:
        works = rec.get("works") or []
        if rec.get("enrich_status") != "ok" or not works:
            continue

        # Annotate EVERY paper once (see annotate_works docstring for why this
        # used to only happen transiently for the 3 picked papers). Persisting
        # this back onto rec["works"] is what lets merge_to_excel.py's
        # "论文明细" sheet — and any notebook cell that reads works straight
        # out of this script's output — show a JCR quartile for EVERY paper,
        # not just the picked representative ones.
        annotated_all = annotate_works(works, jcr_lookup, a.field_major)
        rec["works"] = annotated_all

        works_a_raw = [w for w in annotated_all if (w.get("year") or 0) >= a.window_a_since
                      and not is_ssrn(w) and w.get("is_published")]
        works_b_raw = [w for w in annotated_all if (w.get("year") or 0) >= a.window_b_since]

        for w in works_a_raw:
            match_stats[w["jcr_matched_by"] if w["jcr_matched"] else "unmatched"] += 1

        # 研究方向信号必须在分档/推荐之前算好——它的"排除某个学科桶"这个结果，
        # 要同时应用到窗口A（分档判断）和窗口B（代表作候选），不能只影响其中一处，
        # 否则会出现"这篇论文被判定不可信，但换个窗口又被采信了"这种自相矛盾。
        signal = resolve_domain_signal(
            works_a_raw,
            institution_count=len(rec.get("affiliation_institutions") or []) or None,
            field_hint=a.field_major)

        if signal["block"]:
            rec["journal_ranking"] = {
                "tier": "contaminated", "tier_reason": signal["block_reason"],
                "tier_detail": {"reason": signal["block_reason"]},
                "domain_signal": signal, "pick_reason": signal["block_reason"],
                "window_a_paper_count": len(works_a_raw), "representative_papers": [],
            }
            n_scored += 1
            continue

        excluded = set(signal["excluded_buckets"])
        works_a = ([w for w in works_a_raw if _domain_bucket(w.get("jcr_category")) not in excluded]
                  if excluded else works_a_raw)
        works_b = ([w for w in works_b_raw if _domain_bucket(w.get("jcr_category")) not in excluded]
                  if excluded else works_b_raw)

        tier, detail = classify_professor(works_a, thresholds, a.interests)
        if signal["note"]:
            detail["reason"] = f"（{signal['note']}）{detail['reason']}"
        picked, pick_reason = pick_representative_papers(works_b, tier, a.interests)
        rec["journal_ranking"] = {
            "tier": tier, "tier_reason": detail["reason"], "tier_detail": detail,
            "domain_signal": signal,
            "pick_reason": pick_reason,
            "window_a_paper_count": len(works_a),
            "representative_papers": [
                {"title": w.get("title"), "year": w.get("year"), "venue": w.get("venue"),
                 "url": w.get("url"), "is_ssrn": is_ssrn(w), "coauthors": w.get("coauthors"),
                 "jcr_matched": w.get("jcr_matched"), "jcr_matched_by": w.get("jcr_matched_by"),
                 "jcr_matched_name": w.get("jcr_matched_name"),
                 "jcr_percentile_effective": w.get("jcr_percentile"),
                 "jcr_quartile_effective": w.get("jcr_quartile_effective"),
                 "jcr_index_type": w.get("jcr_index_type"), "jcr_jif": w.get("jcr_jif"),
                 "jcr_rank": w.get("jcr_rank"), "jcr_rank_total": w.get("jcr_rank_total"),
                 "priority_reason": w.get("priority_reason") if not is_ssrn(w) else "SSRN工作论文，按①档规则优先于期刊论文展示",
                 "is_utd24": w.get("is_utd24"), "relevance_note": w.get("relevance_note"),
                 "note": w.get("ssrn_note")}
                for w in picked],
        }
        n_scored += 1

    json.dump(data, open(a.out, "w", encoding="utf-8"), ensure_ascii=False, indent=2)

    total_matched = sum(match_stats.values())
    print(f"\n期刊名匹配情况（窗口A期刊论文，共{total_matched}篇）：", file=sys.stderr)
    for k in ("exact_fullname", "exact_abbr", "exact_issn", "fuzzy", "unmatched"):
        if match_stats.get(k):
            print(f"  {k}: {match_stats[k]}", file=sys.stderr)

    blocked = [r for r in data if (r.get("journal_ranking") or {}).get("tier") == "contaminated"]
    if blocked:
        print(f"\n⛔ {len(blocked)} 位教授的研究方向组合本身不太可能成立，已跳过分档与代表作推荐，"
              f"建议人工核实身份：", file=sys.stderr)
        for r in blocked:
            print(f"  - {r.get('name')} ({r.get('institution')})：{r['journal_ranking']['tier_reason']}",
                  file=sys.stderr)

    noted = [r for r in data if (r.get("journal_ranking") or {}).get("domain_signal", {}).get("note")]
    if noted:
        print(f"\n💡 {len(noted)} 位教授窗口A论文里有研究方向提示（仅供参考，代表作推荐照常进行）：",
              file=sys.stderr)
        for r in noted:
            print(f"  - {r.get('name')} ({r.get('institution')})：{r['journal_ranking']['domain_signal']['note']}",
                  file=sys.stderr)

    tiers = Counter(r["journal_ranking"]["tier"] for r in data if r.get("journal_ranking"))
    print(f"\n分档结果（{n_scored} 位已成功识别身份的教授）：{dict(tiers)}", file=sys.stderr)

    print(f"wrote {a.out}", file=sys.stderr)


if __name__ == "__main__":
    main()
